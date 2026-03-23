from fastapi import FastAPI, Depends, HTTPException, status, Request, UploadFile, File, Form, Header
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import text
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from jose import JWTError, jwt
import json
import os
import calendar
import smtplib
import base64
import hashlib
from uuid import uuid4
from collections import defaultdict
from email.message import EmailMessage
from urllib.parse import urlencode
from urllib.request import Request as UrlRequest, urlopen
from urllib.error import HTTPError, URLError
import ssl
import time
import resource
import asyncio
import threading
from fastapi.responses import HTMLResponse, PlainTextResponse
from pypdf import PdfReader
from openpyxl import load_workbook
import xlrd

from database import get_db, init_db, engine
from models import User, Transaction, Asset, Document, Budget, Goal, Investment, Liability, Integration, AuditLog, AppInsightsMetric, Base, UserRole
from schemas import (
    UserCreate, User as UserSchema, Transaction as TransactionSchema,
    Asset as AssetSchema, AssetCreate, AssetUpdate,
    Document as DocumentSchema, DocumentUpdate,
    TransactionCreate, TransactionUpdate, Budget as BudgetSchema,
    BudgetCreate, BudgetUpdate, BudgetWithSpending, Goal as GoalSchema, GoalCreate, GoalUpdate,
    Investment as InvestmentSchema, InvestmentCreate, InvestmentUpdate,
    Liability as LiabilitySchema, LiabilityCreate, LiabilityUpdate,
    Integration as IntegrationSchema, IntegrationCreate, IntegrationUpdate,
    AuditLog as AuditLogSchema, Token, DashboardStats, TransactionFilter,
    AuditLogFilter, LoginRequest, UserRoleUpdate, UserStatusUpdate, UserPermissionsUpdate,
    UserProfileUpdate, UserPasswordChange,
    ExternalRBACProvisionRequest, FederatedClaimSyncRequest,
    AdminResetRequest, AdminResetResponse,
    ExpenseReportEmailRequest, ExpenseReportEmailResponse, IntegrationAuthUrlResponse,
    DataImportResponse, GmailBankAlertSyncResult, StatementImportResponse,
    StatementImportTransaction
)
from auth import (
    get_password_hash, verify_password, authenticate_user, create_access_token,
    get_current_active_user, ACCESS_TOKEN_EXPIRE_MINUTES,
    get_client_ip, check_registration_rate_limit, require_admin, require_superadmin, require_write_access,
    SECRET_KEY, ALGORITHM, SUPERADMIN_USERNAME, SUPERADMIN_PASSWORD,
    get_default_permissions, normalize_permissions, user_has_permission
)
from integration_providers import get_provider

import csv
import io
import re

from models import TransactionType

# Create tables and ensure migrations
init_db()

app = FastAPI(title="Ledger Finance API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify exact origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static upload directory
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount('/uploads', StaticFiles(directory=UPLOAD_DIR), name='uploads')


APP_INSIGHTS_STATE = {
    "started_at": datetime.utcnow(),
    "total_requests": 0,
    "active_requests": 0,
    "errors_5xx": 0,
    "discards_4xx": 0,
    "backend_latency_ms_total": 0.0,
    "backend_latency_samples": 0,
    "last_errors": [],
}
APP_INSIGHTS_HISTORY: List[Dict] = []
MAX_INSIGHTS_HISTORY = 60


def _record_app_error(path: str, status_code: int, elapsed_ms: float) -> None:
    APP_INSIGHTS_STATE["last_errors"].append({
        "time": datetime.utcnow().isoformat(),
        "path": path,
        "status": status_code,
        "latency_ms": round(elapsed_ms, 2),
    })
    # Keep only recent errors to avoid unbounded growth.
    if len(APP_INSIGHTS_STATE["last_errors"]) > 30:
        APP_INSIGHTS_STATE["last_errors"] = APP_INSIGHTS_STATE["last_errors"][-30:]


@app.middleware("http")
async def collect_runtime_metrics(request: Request, call_next):
    start = time.perf_counter()
    APP_INSIGHTS_STATE["total_requests"] += 1
    APP_INSIGHTS_STATE["active_requests"] += 1

    try:
        response = await call_next(request)
        elapsed_ms = (time.perf_counter() - start) * 1000.0
        APP_INSIGHTS_STATE["backend_latency_ms_total"] += elapsed_ms
        APP_INSIGHTS_STATE["backend_latency_samples"] += 1

        if response.status_code >= 500:
            APP_INSIGHTS_STATE["errors_5xx"] += 1
            _record_app_error(request.url.path, response.status_code, elapsed_ms)
        elif response.status_code >= 400:
            APP_INSIGHTS_STATE["discards_4xx"] += 1

        return response
    except Exception:
        elapsed_ms = (time.perf_counter() - start) * 1000.0
        APP_INSIGHTS_STATE["backend_latency_ms_total"] += elapsed_ms
        APP_INSIGHTS_STATE["backend_latency_samples"] += 1
        APP_INSIGHTS_STATE["errors_5xx"] += 1
        _record_app_error(request.url.path, 500, elapsed_ms)
        raise
    finally:
        APP_INSIGHTS_STATE["active_requests"] = max(0, APP_INSIGHTS_STATE["active_requests"] - 1)


def _get_mysql_status_map(db: Session, variables: List[str]) -> Dict[str, int]:
    status_map: Dict[str, int] = {}
    for var in variables:
        safe_var = var.replace("'", "")
        row = db.execute(text(f"SHOW GLOBAL STATUS LIKE '{safe_var}'")).fetchone()
        value = 0
        if row and len(row) >= 2:
            try:
                value = int(float(row[1]))
            except (TypeError, ValueError):
                value = 0
        status_map[var] = value
    return status_map


def _fetch_frontend_nginx_status() -> Dict[str, Optional[float]]:
    # In docker-compose the frontend service is reachable as "frontend".
    target_url = os.getenv("FRONTEND_STATUS_URL", "https://frontend/nginx_status")
    context = ssl._create_unverified_context()

    try:
        request = UrlRequest(target_url, headers={"User-Agent": "ledger-backend-app-insights"})
        with urlopen(request, timeout=3, context=context) as response:
            payload = response.read().decode("utf-8", errors="ignore")
    except Exception:
        return {
            "status": "unavailable",
            "connection_count": None,
            "errors": None,
            "discards": None,
            "reading": None,
            "writing": None,
            "waiting": None,
            "requests_total": None,
        }

    lines = [line.strip() for line in payload.splitlines() if line.strip()]
    active_connections = None
    accepts = None
    handled = None
    requests_total = None
    reading = None
    writing = None
    waiting = None

    # Expected format:
    # Active connections: 3
    # server accepts handled requests
    #  8 8 30
    # Reading: 0 Writing: 1 Waiting: 2
    for idx, line in enumerate(lines):
        if line.lower().startswith("active connections"):
            try:
                active_connections = int(line.split(":", 1)[1].strip())
            except (ValueError, IndexError):
                active_connections = None
        elif line.lower().startswith("server accepts handled requests") and idx + 1 < len(lines):
            parts = lines[idx + 1].split()
            if len(parts) >= 3:
                try:
                    accepts = int(parts[0])
                    handled = int(parts[1])
                    requests_total = int(parts[2])
                except ValueError:
                    accepts = handled = requests_total = None
        elif line.lower().startswith("reading:"):
            try:
                # Reading: 0 Writing: 1 Waiting: 2
                parts = line.replace(":", "").split()
                reading = int(parts[1])
                writing = int(parts[3])
                waiting = int(parts[5])
            except (ValueError, IndexError):
                reading = writing = waiting = None

    return {
        "status": "ok" if active_connections is not None else "degraded",
        "connection_count": active_connections,
        # Nginx stub_status doesn't provide direct error/discard counters.
        "errors": None,
        "discards": None,
        "reading": reading,
        "writing": writing,
        "waiting": waiting,
        "accepts": accepts,
        "handled": handled,
        "requests_total": requests_total,
    }


def _safe_iso_ts(value: Optional[str], fallback: Optional[datetime] = None) -> str:
    if isinstance(value, str) and value.strip():
        return value
    return (fallback or datetime.utcnow()).isoformat()


def serialize_user(user: User) -> Dict:
    permissions = normalize_permissions(user.role, getattr(user, "permissions_json", None))
    return {
        "id": user.id,
        "email": user.email,
        "username": user.username,
        "full_name": user.full_name,
        "role": user.role,
        "permissions": permissions,
        "identity_provider": getattr(user, "identity_provider", None),
        "external_subject": getattr(user, "external_subject", None),
        "is_active": user.is_active,
        "created_at": user.created_at,
    }


def ensure_superadmin_account() -> None:
    db = next(get_db())
    try:
        superadmin = db.query(User).filter(User.username == SUPERADMIN_USERNAME).first()
        superadmin_permissions = json.dumps(get_default_permissions(UserRole.SUPERADMIN.value))

        if not superadmin:
            superadmin = User(
                email=f"{SUPERADMIN_USERNAME}@local.ledger-app",
                username=SUPERADMIN_USERNAME,
                full_name="Super Admin",
                hashed_password=get_password_hash(SUPERADMIN_PASSWORD),
                role=UserRole.SUPERADMIN.value,
                permissions_json=superadmin_permissions,
                is_active=True,
            )
            db.add(superadmin)
            db.commit()
            db.refresh(superadmin)
            log_audit(db, superadmin.id, "create", "user", str(superadmin.id), "Auto-provisioned default super admin account")
            return

        changed = False
        if superadmin.role != UserRole.SUPERADMIN.value:
            superadmin.role = UserRole.SUPERADMIN.value
            changed = True
        if superadmin.permissions_json != superadmin_permissions:
            superadmin.permissions_json = superadmin_permissions
            changed = True
        if not superadmin.is_active:
            superadmin.is_active = True
            changed = True
        if changed:
            db.commit()
    finally:
        db.close()


EXTERNAL_RBAC_API_KEY = os.getenv("EXTERNAL_RBAC_API_KEY", "").strip()


def verify_external_rbac_api_key(authorization: Optional[str], x_rbac_api_key: Optional[str]) -> None:
    """Validate machine-to-machine RBAC provisioning key for external IdP systems."""
    configured_key = EXTERNAL_RBAC_API_KEY
    if not configured_key:
        raise HTTPException(status_code=503, detail="External RBAC API is disabled. Configure EXTERNAL_RBAC_API_KEY.")

    provided_key = (x_rbac_api_key or "").strip()
    if not provided_key and authorization:
        auth_value = authorization.strip()
        if auth_value.lower().startswith("bearer "):
            provided_key = auth_value[7:].strip()

    if not provided_key or provided_key != configured_key:
        raise HTTPException(status_code=403, detail="Invalid external RBAC API key")


def resolve_federated_role(groups: List[str], role_hint: Optional[UserRole]) -> str:
    normalized_groups = {str(group).strip().lower() for group in (groups or []) if group is not None}
    if role_hint == UserRole.SUPERADMIN or "superadmin" in normalized_groups:
        return UserRole.SUPERADMIN.value
    if role_hint == UserRole.ADMIN or "admin" in normalized_groups:
        return UserRole.ADMIN.value
    if role_hint == UserRole.READONLY or "readonly" in normalized_groups or "read_only" in normalized_groups:
        return UserRole.READONLY.value
    if role_hint == UserRole.USER:
        return UserRole.USER.value
    return UserRole.USER.value


def _derive_instance_logs(
    history: List[Dict],
    frontend_instance: Dict,
    backend_instance: Dict,
    db_instance: Dict,
    alerts: List[Dict],
    backend_recent_errors: Optional[List[Dict]] = None,
) -> Dict[str, List[Dict[str, str]]]:
    logs: Dict[str, List[Dict[str, str]]] = {
        "frontend": [],
        "backend": [],
        "db": [],
    }

    def add_log(instance: str, level: str, message: str, ts: Optional[str] = None, source: str = "metrics") -> None:
        if instance not in logs:
            return
        logs[instance].append({
            "ts": _safe_iso_ts(ts),
            "level": level,
            "source": source,
            "message": message,
        })

    # Backend logs: concrete recent errors + signal-based warnings from snapshots.
    for err in (backend_recent_errors or [])[-30:]:
        path = err.get("path") or "unknown"
        status_code = err.get("status") or "—"
        latency = err.get("latency_ms")
        latency_txt = f" · {latency} ms" if latency is not None else ""
        add_log(
            "backend",
            "error",
            f"{status_code} on {path}{latency_txt}",
            ts=err.get("time"),
            source="http",
        )

    for snap in history[-80:]:
        ts = snap.get("ts")
        be_err = int(snap.get("backend_errors") or 0)
        be_4xx = int(snap.get("backend_discards") or 0)
        be_lat = float(snap.get("backend_latency_ms") or 0)
        if be_err > 0:
            add_log("backend", "error", f"Backend 5xx total observed: {be_err}", ts=ts, source="timeseries")
        if be_4xx > 50:
            add_log("backend", "warn", f"Backend 4xx/discards elevated: {be_4xx}", ts=ts, source="timeseries")
        if be_lat >= 500:
            add_log("backend", "warn", f"Backend latency high: {round(be_lat, 2)} ms", ts=ts, source="timeseries")

        fe_conn = snap.get("frontend_connections")
        fe_wait = snap.get("frontend_waiting")
        if fe_conn is not None and fe_conn > 300:
            add_log("frontend", "warn", f"Frontend active connections high: {fe_conn}", ts=ts, source="timeseries")
        if fe_wait is not None and fe_wait > 200:
            add_log("frontend", "warn", f"Frontend waiting connections elevated: {fe_wait}", ts=ts, source="timeseries")

        db_conn = int(snap.get("db_connections") or 0)
        db_err = int(snap.get("db_errors") or 0)
        db_slow = int(snap.get("db_slow_queries") or 0)
        if db_conn > 80:
            add_log("db", "warn", f"DB connections high: {db_conn}", ts=ts, source="timeseries")
        if db_err > 0:
            add_log("db", "error", f"DB connection/client errors observed: {db_err}", ts=ts, source="timeseries")
        if db_slow > 0:
            add_log("db", "warn", f"DB slow queries observed: {db_slow}", ts=ts, source="timeseries")

    # Alert feed -> instance-specific logs.
    severity_to_level = {"high": "error", "medium": "warn", "low": "info"}
    for alert in alerts:
        instance = (alert.get("instance") or "").lower()
        if instance not in logs:
            continue
        level = severity_to_level.get((alert.get("severity") or "").lower(), "info")
        add_log(instance, level, alert.get("message") or "Alert", source="alert")

    now_ts = datetime.utcnow().isoformat()
    if frontend_instance.get("status") == "ok":
        add_log(
            "frontend",
            "info",
            f"Frontend healthy · active={frontend_instance.get('connection_count')}, waiting={frontend_instance.get('waiting')}",
            ts=now_ts,
            source="health",
        )
    if backend_instance.get("status") == "ok":
        add_log(
            "backend",
            "info",
            f"Backend healthy · active={backend_instance.get('connection_count')}, latency={backend_instance.get('details', {}).get('avg_latency_ms')} ms",
            ts=now_ts,
            source="health",
        )
    if db_instance.get("status") == "ok":
        add_log(
            "db",
            "info",
            f"DB healthy · connections={db_instance.get('connection_count')}, threads={db_instance.get('details', {}).get('threads_running')}",
            ts=now_ts,
            source="health",
        )

    # Sort descending by timestamp and cap each instance log stream.
    for instance in logs:
        logs[instance] = sorted(logs[instance], key=lambda item: item.get("ts", ""), reverse=True)[:250]

    return logs


# Helper function to log audit
def log_audit(db: Session, user_id: int, action: str, entity_type: str, 
               entity_id: str, description: str, details: dict = None, 
               user_agent: str = None):
    """
    Log audit trail entry
    Handles datetime serialization for JSON
    """
    # Convert details to JSON-serializable format
    if details:
        import datetime

        def make_serializable(obj):
            if isinstance(obj, (datetime.datetime, datetime.date)):
                return obj.isoformat()
            if isinstance(obj, dict):
                return {k: make_serializable(v) for k, v in obj.items()}
            if isinstance(obj, list):
                return [make_serializable(v) for v in obj]
            return obj

        serializable_details = make_serializable(details)
        details_json = json.dumps(serializable_details)
    else:
        details_json = None
    
    audit_log = AuditLog(
        user_id=user_id,
        action=action,
        entity_type=entity_type,
        entity_id=str(entity_id),
        description=description,
        details=details_json,
        user_agent=user_agent
    )
    db.add(audit_log)
    db.commit()


def _serialize_document(document: Document) -> DocumentSchema:
    return DocumentSchema(
        id=document.id,
        user_id=document.user_id,
        title=document.title,
        folder=document.folder or "General",
        subfolder=document.subfolder,
        file_name=document.file_name,
        content_type=document.content_type,
        document_type=document.document_type or "general",
        frozen_import=bool(document.frozen_import),
        imported_transaction_count=int(document.imported_transaction_count or 0),
        uploaded_at=document.uploaded_at,
        url=f"/uploads/{os.path.basename(document.file_path)}"
    )


def _statement_document_note(document_id: int) -> str:
    return f"statement_doc_id:{document_id}"


STATEMENT_SALARY_NEXT_MONTH_TAG = "salary_credit_next_month"
STATEMENT_NON_INCOME_CREDIT_TAG = "statement_non_income_credit"
STATEMENT_BALANCE_RECONCILED_TAG = "statement_balance_reconciled"
STATEMENT_BALANCE_TOLERANCE = 1.0


STATEMENT_CREDIT_ADJUSTMENT_KEYWORDS = {
    "reversal",
    "reversed",
    "refund",
    "refunded",
    "chargeback",
    "cashback",
    "returned",
    "return credit",
    "txn rev",
    "txn reversal",
    "failed txn",
    "failed transaction",
}


STATEMENT_SKIP_KEYWORDS = {
    "opening balance",
    "closing balance",
    "available balance",
    "statement summary",
    "total amount due",
    "minimum amount due",
    "payment due date",
    "credit limit",
    "transaction details",
    "transaction date",
    "value date",
    "description",
    "particulars",
    "reward points",
    "customer care",
    "branch",
    "ifsc",
    "page ",
}

STATEMENT_CATEGORY_KEYWORDS = {
    "Investment": [" sip ", "systematic investment", "mutual fund", "zerodha", "groww", "nps", "ppf", "coin by zerodha"],
    "Loans (EMI)": [" emi ", "loan", "installment", "nach", "ecs", "finance", "home loan", "personal loan", "car loan"],
    "Food": ["swiggy", "zomato", "restaurant", "cafe", "domino", "pizza", "eatery"],
    "Travel": ["uber", "ola", "rapido", "irctc", "metro", "flight", "train", "bus", "fuel", "petrol"],
    "Shopping": ["amazon", "flipkart", "myntra", "ajio", "store purchase", "shopping"],
    "Bills": ["electricity", "water", "gas", "internet", "broadband", "airtel", "jio", "recharge", "utility"],
    "Health": ["hospital", "clinic", "doctor", "pharmacy", "medical"],
    "Life Insurance": ["life insurance", "term insurance", "lic", "policy premium"],
    "Health Insurance": ["health insurance", "mediclaim"],
    "Motor Insurance": ["motor insurance", "car insurance", "bike insurance"],
    "Salary": ["salary", " sal ", "sal/", "payroll", "pay credit", "salary credit", "salary cr", "wage credit", "salary transfer"],
    "Rental Income": ["rent received", "rental income"],
    "Business": ["invoice payment", "client payment", "business receipt"],
    "Refund / Reversal": ["reversal", "refund", "chargeback", "cashback", "returned", "return credit", "failed txn"],
}

STATEMENT_TABLE_DATE_REGEX = r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}"
STATEMENT_FILE_EXTENSIONS = {".csv", ".xls", ".xlsx"}
STATEMENT_HEADER_ALIASES = {
    "date": "date",
    "narration": "narration",
    "chqrefno": "reference",
    "chqrefnumber": "reference",
    "refno": "reference",
    "valuedt": "value_date",
    "value date": "value_date",
    "withdrawalamt": "withdrawal",
    "withdrawal amt": "withdrawal",
    "depositamt": "deposit",
    "deposit amt": "deposit",
    "closingbalance": "closing_balance",
    "closing balance": "closing_balance",
}


def _extract_statement_pdf_text(file_bytes: bytes, password: Optional[str] = None) -> str:
    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        if reader.is_encrypted:
            if not password:
                raise HTTPException(status_code=400, detail="PDF_PASSWORD_REQUIRED")

            decrypt_result = reader.decrypt(password)
            if decrypt_result == 0:
                raise HTTPException(status_code=400, detail="INVALID_PDF_PASSWORD")

        text = "\n".join((page.extract_text() or "") for page in reader.pages)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Unable to read PDF statement: {exc}") from exc

    if not text.strip():
        raise HTTPException(status_code=400, detail="PDF has no extractable text. Scanned image PDFs are not supported yet.")

    return text


def _parse_statement_date(value: str) -> Optional[datetime]:
    candidates = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d %b %Y",
        "%d %b %y",
        "%d %B %Y",
        "%d %B %y",
        "%d %b",
        "%d %B",
    ]
    for fmt in candidates:
        try:
            parsed = datetime.strptime(value, fmt)
            if "%Y" not in fmt and "%y" not in fmt:
                now = datetime.utcnow()
                parsed = parsed.replace(year=now.year)
            return parsed
        except ValueError:
            continue
    return None


def _is_statement_credit_adjustment(description: str) -> bool:
    normalized = f" {description.lower()} "
    return any(keyword in normalized for keyword in STATEMENT_CREDIT_ADJUSTMENT_KEYWORDS)


def _reconcile_statement_balance_rows(rows: List[Dict[str, object]]) -> None:
    ascending_matches = 0
    descending_matches = 0

    for index in range(1, len(rows)):
        previous_row = rows[index - 1]
        current_row = rows[index]

        previous_balance = previous_row.get("closing_balance")
        current_balance = current_row.get("closing_balance")
        if previous_balance is None or current_balance is None:
            continue

        net_flow = float(current_row.get("deposit_amt") or 0) - float(current_row.get("withdrawal_amt") or 0)
        if abs(net_flow) <= STATEMENT_BALANCE_TOLERANCE:
            continue

        ascending_delta = float(current_balance) - float(previous_balance)
        descending_delta = float(previous_balance) - float(current_balance)

        if abs(ascending_delta - net_flow) <= STATEMENT_BALANCE_TOLERANCE:
            ascending_matches += 1
        if abs(descending_delta - net_flow) <= STATEMENT_BALANCE_TOLERANCE:
            descending_matches += 1

    statement_order = 0
    if ascending_matches > 0 or descending_matches > 0:
        statement_order = 1 if ascending_matches >= descending_matches else -1

    for index, row in enumerate(rows):
        row["balance_delta"] = None
        row["balance_reconciled"] = False
        if statement_order == 0 or index == 0:
            continue

        previous_row = rows[index - 1]
        previous_balance = previous_row.get("closing_balance")
        current_balance = row.get("closing_balance")
        if previous_balance is None or current_balance is None:
            continue

        balance_delta = (
            float(current_balance) - float(previous_balance)
            if statement_order > 0
            else float(previous_balance) - float(current_balance)
        )
        expected_delta = float(row.get("deposit_amt") or 0) - float(row.get("withdrawal_amt") or 0)

        row["balance_delta"] = balance_delta
        row["balance_reconciled"] = abs(balance_delta - expected_delta) <= STATEMENT_BALANCE_TOLERANCE


def _extract_statement_amounts(text: str) -> List[float]:
    amounts = []
    # Prefer decimal amounts to avoid accidentally treating long ref numbers as money.
    matches = re.findall(r"([+-]?\d[\d,]*\.\d{1,2})", text, flags=re.IGNORECASE)

    # Fallback for integer amounts only when clearly marked by currency symbols.
    if not matches:
        matches = re.findall(r"(?:INR|Rs\.?|₹)\s*([+-]?\d[\d,]*(?:\.\d{1,2})?)", text, flags=re.IGNORECASE)

    for raw_amount in matches:
        try:
            amounts.append(abs(float(raw_amount.replace(',', ''))))
        except ValueError:
            continue
    return amounts


def _clean_statement_description(text: str) -> str:
    cleaned = re.sub(r"(?:INR|Rs\.?|₹)?\s*[+-]?\d[\d,]*(?:\.\d{1,2})?(?:\s*(?:CR|DR))?", " ", text, flags=re.IGNORECASE)
    cleaned = re.sub(r"\b(?:CR|DR|CREDIT|DEBIT)\b", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -|")
    return cleaned[:240]


def _classify_statement_transaction(description: str, full_line: str) -> Tuple[str, str]:
    normalized = f" {description.lower()} {full_line.lower()} "

    category = "Other"
    for candidate, keywords in STATEMENT_CATEGORY_KEYWORDS.items():
        if any(keyword in normalized for keyword in keywords):
            category = candidate
            break

    income_hints = [
        " credit ", " cr ", " salary ", " salary credit ", " salary cr ",
        " refund ", " cashback ", " interest ", " received ", " inward ",
        " neft in ", " imps in ", " rtgs in ", " transfer from "
    ]
    expense_hints = [" debit ", " dr ", " purchase ", " pos ", " upi ", " bill ", " payment ", " nach ", " ecs ", " transfer to "]
    is_income = any(hint in normalized for hint in income_hints)
    is_expense = any(hint in normalized for hint in expense_hints)
    has_credit_marker = bool(re.search(r"\bcr\b|\bcredit(?:ed)?\b", normalized, re.IGNORECASE))
    has_debit_marker = bool(re.search(r"\bdr\b|\bdebit(?:ed)?\b", normalized, re.IGNORECASE))
    salary_marker = bool(re.search(r"\bsalary\b|\bsal(?:ary)?\b|\bslry\b|\bpayroll\b", normalized, re.IGNORECASE))

    if salary_marker:
        return "income", "Salary"

    if category in {"Salary", "Rental Income", "Business"}:
        return "income", category
    if has_credit_marker and not has_debit_marker:
        return "income", category if category != "Other" else "Other Sources"
    if has_debit_marker and not has_credit_marker:
        return "expense", category
    if is_income and not is_expense:
        return "income", category if category != "Other" else "Other Sources"
    return "expense", category


def _parse_statement_table_row(line: str) -> Optional[Dict[str, object]]:
    """
    Parse rows shaped like:
    Date Narration Chq./Ref.No. Value Dt Withdrawal Amt. Deposit Amt. Closing Balance
    """
    date_prefix_match = re.match(rf"^({STATEMENT_TABLE_DATE_REGEX})\s+", line)
    if not date_prefix_match:
        return None

    txn_date = _parse_statement_date(date_prefix_match.group(1).strip())
    if txn_date is None:
        return None

    remaining = line[date_prefix_match.end():].strip()
    value_date_match = re.search(rf"\b({STATEMENT_TABLE_DATE_REGEX})\b", remaining)
    if not value_date_match:
        return None

    narration_with_ref = remaining[:value_date_match.start()].strip()
    post_value_date = remaining[value_date_match.end():].strip()

    amount_tokens = re.findall(r"([+-]?\d[\d,]*\.\d{1,2})", post_value_date)
    if len(amount_tokens) < 3:
        return None

    try:
        withdrawal_amt = abs(float(amount_tokens[0].replace(',', '')))
        deposit_amt = abs(float(amount_tokens[1].replace(',', '')))
    except ValueError:
        return None

    if withdrawal_amt <= 0 and deposit_amt <= 0:
        return None

    description = _clean_statement_description(narration_with_ref)
    if not description:
        description = "Statement Transaction"

    tx_type, category = _classify_statement_transaction(description, line)

    # Column-driven mapping wins for this statement format.
    if withdrawal_amt > 0 and deposit_amt <= 0:
        tx_type = "expense"
        amount = withdrawal_amt
    elif deposit_amt > 0 and withdrawal_amt <= 0:
        tx_type = "income"
        amount = deposit_amt
    else:
        amount = max(withdrawal_amt, deposit_amt)

    # Extra-safe salary handling for uppercase SALARY/SAL/SLRY variants in narration.
    salary_marker = bool(re.search(r"\bsalary\b|\bsal(?:ary)?\b|\bslry\b|\bpayroll\b", f" {description.lower()} {line.lower()} "))
    if salary_marker:
        tx_type = "income"
        category = "Salary"
        if deposit_amt > 0:
            amount = deposit_amt

    if amount <= 0:
        return None

    return {
        "date": txn_date,
        "description": description,
        "amount": amount,
        "type": tx_type,
        "category": category,
    }


def _normalize_statement_header(value: object) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    compact = re.sub(r"[^a-z0-9]+", "", raw)
    return STATEMENT_HEADER_ALIASES.get(compact, STATEMENT_HEADER_ALIASES.get(raw, compact))


def _parse_statement_amount_cell(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return abs(float(value))

    raw = str(value).strip()
    if not raw:
        return 0.0

    cleaned = re.sub(r"[^0-9.\-]", "", raw.replace(',', ''))
    if not cleaned or cleaned in {"-", ".", "-."}:
        return 0.0

    try:
        return abs(float(cleaned))
    except ValueError:
        return 0.0


def _parse_statement_date_cell(value: object) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(hour=0, minute=0, second=0, microsecond=0)
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return datetime(value.year, value.month, value.day)

    raw = str(value).strip()
    if not raw:
        return None
    return _parse_statement_date(raw)


def _read_statement_sheet_rows(file_name: str, file_bytes: bytes) -> List[List[object]]:
    extension = os.path.splitext(file_name or "")[1].lower()
    if extension == ".csv":
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")
        return [row for row in csv.reader(io.StringIO(text))]

    if extension == ".xlsx":
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    if extension == ".xls":
        workbook = xlrd.open_workbook(file_contents=file_bytes)
        sheet = workbook.sheet_by_index(0)
        rows: List[List[object]] = []
        for row_idx in range(sheet.nrows):
            row_values: List[object] = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    dt_value = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
                    row_values.append(dt_value)
                else:
                    row_values.append(cell.value)
            rows.append(row_values)
        return rows

    raise HTTPException(status_code=400, detail="Please upload a CSV, XLS, or XLSX bank statement")


def _find_statement_header_map(rows: List[List[object]]) -> Tuple[int, Dict[str, int]]:
    required = {"date", "narration", "withdrawal", "deposit"}

    for row_index, row in enumerate(rows[:25]):
        header_map: Dict[str, int] = {}
        for col_index, cell_value in enumerate(row):
            normalized = _normalize_statement_header(cell_value)
            if normalized and normalized not in header_map:
                header_map[normalized] = col_index

        if required.issubset(header_map.keys()):
            return row_index, header_map

    raise HTTPException(
        status_code=400,
        detail="Statement file must contain headers for Date, Narration, Withdrawal Amt, and Deposit Amt"
    )


def _parse_statement_spreadsheet_transactions(file_name: str, file_bytes: bytes) -> List[Dict[str, object]]:
    rows = _read_statement_sheet_rows(file_name, file_bytes)
    if not rows:
        raise HTTPException(status_code=400, detail="Uploaded statement file is empty")

    header_row_index, header_map = _find_statement_header_map(rows)
    parsed_rows: List[Dict[str, object]] = []

    for row in rows[header_row_index + 1:]:
        if not row or not any(str(cell).strip() for cell in row if cell is not None):
            continue

        def get_cell(field_name: str) -> object:
            index = header_map.get(field_name)
            if index is None or index >= len(row):
                return None
            return row[index]

        transaction_date = _parse_statement_date_cell(get_cell("date")) or _parse_statement_date_cell(get_cell("value_date"))
        if transaction_date is None:
            continue

        narration = str(get_cell("narration") or "").strip()
        reference = str(get_cell("reference") or "").strip()
        description = " ".join(part for part in [narration, reference] if part).strip()
        if not description:
            continue

        withdrawal_amt = _parse_statement_amount_cell(get_cell("withdrawal"))
        deposit_amt = _parse_statement_amount_cell(get_cell("deposit"))
        if withdrawal_amt <= 0 and deposit_amt <= 0:
            continue

        closing_balance = _parse_statement_amount_cell(get_cell("closing_balance"))
        if closing_balance <= 0:
            closing_balance = None

        parsed_rows.append({
            "date": transaction_date,
            "description": _clean_statement_description(description),
            "withdrawal_amt": withdrawal_amt,
            "deposit_amt": deposit_amt,
            "closing_balance": closing_balance,
        })

    _reconcile_statement_balance_rows(parsed_rows)

    parsed_transactions: List[Dict[str, object]] = []
    for row in parsed_rows:
        description = str(row.get("description") or "").strip()
        if not description:
            continue

        withdrawal_amt = float(row.get("withdrawal_amt") or 0)
        deposit_amt = float(row.get("deposit_amt") or 0)
        balance_delta = row.get("balance_delta")

        tx_type, category = _classify_statement_transaction(description, description)
        if isinstance(balance_delta, (int, float)) and abs(float(balance_delta)) > STATEMENT_BALANCE_TOLERANCE:
            tx_type = "income" if float(balance_delta) > 0 else "expense"
            amount = abs(float(balance_delta))
        elif withdrawal_amt > 0 and deposit_amt <= 0:
            tx_type = "expense"
            amount = withdrawal_amt
        elif deposit_amt > 0 and withdrawal_amt <= 0:
            tx_type = "income"
            amount = deposit_amt
        else:
            amount = max(withdrawal_amt, deposit_amt)

        salary_marker = bool(re.search(r"\bsalary\b|\bsal(?:ary)?\b|\bslry\b|\bpayroll\b", f" {description.lower()} ", re.IGNORECASE))
        if salary_marker and tx_type == "income":
            tx_type = "income"
            category = "Salary"
            if deposit_amt > 0:
                amount = deposit_amt

        is_credit_adjustment = tx_type == "income" and _is_statement_credit_adjustment(description)
        if is_credit_adjustment:
            category = "Refund / Reversal"

        if amount <= 0:
            continue

        parsed_transactions.append({
            "date": row["date"],
            "description": description,
            "amount": amount,
            "type": tx_type,
            "category": category,
            "is_credit_adjustment": is_credit_adjustment,
            "balance_reconciled": bool(row.get("balance_reconciled")),
        })

    unique_transactions: List[Dict[str, object]] = []
    seen_keys = set()
    for entry in parsed_transactions:
        key = (
            entry["date"].date().isoformat(),
            round(float(entry["amount"]), 2),
            str(entry["type"]),
            str(entry["description"]).lower(),
        )
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_transactions.append(entry)

    return unique_transactions


def _parse_statement_transactions(file_bytes: bytes, password: Optional[str] = None) -> List[Dict[str, object]]:
    text = _extract_statement_pdf_text(file_bytes, password=password)
    parsed_transactions = []
    has_withdrawal_deposit_columns = bool(re.search(r"withdrawal\s*amt", text, re.IGNORECASE) and re.search(r"deposit\s*amt", text, re.IGNORECASE))

    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        lowered_line = line.lower()
        if len(line) < 12 or any(keyword in lowered_line for keyword in STATEMENT_SKIP_KEYWORDS):
            continue

        if has_withdrawal_deposit_columns:
            parsed_row = _parse_statement_table_row(line)
            if parsed_row is not None:
                parsed_transactions.append(parsed_row)
                continue

        date_match = re.match(r"^(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3,9}(?:\s+\d{2,4})?)\s+", line)
        if not date_match:
            continue

        transaction_date = _parse_statement_date(date_match.group(1).strip())
        if transaction_date is None:
            continue

        remaining = line[date_match.end():].strip()
        amounts = _extract_statement_amounts(remaining)
        if not amounts:
            continue

        description = _clean_statement_description(remaining)
        if not description:
            continue

        tx_type, category = _classify_statement_transaction(description, line)
        amount = amounts[-2] if len(amounts) >= 2 else amounts[-1]
        salary_detected = bool(re.search(r"\bsalary\b|\bsal(?:ary)?\b|\bslry\b|\bpayroll\b", f" {description.lower()} {line.lower()} ", re.IGNORECASE))

        # For statements with explicit Withdrawal/Deposit columns, map values deterministically.
        if has_withdrawal_deposit_columns and len(amounts) >= 2:
            withdrawal_amt = float(amounts[0])
            deposit_amt = float(amounts[1])
            withdrawal_non_zero = withdrawal_amt > 0.0
            deposit_non_zero = deposit_amt > 0.0

            if withdrawal_non_zero and not deposit_non_zero:
                tx_type = "expense"
                amount = withdrawal_amt
            elif deposit_non_zero and not withdrawal_non_zero:
                tx_type = "income"
                amount = deposit_amt

        if salary_detected:
            tx_type = "income"
            category = "Salary"
            if has_withdrawal_deposit_columns and len(amounts) >= 2:
                # Prefer the deposit column when the statement exposes explicit DR/CR columns.
                deposit_candidate = float(amounts[1])
                if deposit_candidate > 0:
                    amount = deposit_candidate

        if amount <= 0:
            continue

        parsed_transactions.append({
            "date": transaction_date,
            "description": description,
            "amount": amount,
            "type": tx_type,
            "category": category,
        })

    unique_transactions = []
    seen_keys = set()
    for entry in parsed_transactions:
        key = (
            entry["date"].date().isoformat(),
            round(float(entry["amount"]), 2),
            str(entry["type"]),
            str(entry["description"]).lower(),
        )
        if key in seen_keys:
            continue
        seen_keys.add(key)
        unique_transactions.append(entry)

    return unique_transactions


def _statement_fingerprint(entry: Dict[str, object]) -> str:
    raw_value = "|".join([
        str(entry["date"].date().isoformat()),
        f"{float(entry['amount']):.2f}",
        str(entry["type"]),
        str(entry["category"]),
        str(entry["description"]).strip().lower(),
    ])
    return hashlib.sha1(raw_value.encode("utf-8")).hexdigest()[:20]


def _statement_lender_name(description: str) -> str:
    cleaned = re.sub(r"\s+", " ", (description or "")).strip(" -|")
    if not cleaned:
        return "Statement Loan"
    if len(cleaned) > 120:
        cleaned = cleaned[:120].rstrip()
    return cleaned


def _is_statement_document(document: Document) -> bool:
    return (document.document_type or "general") in {"statement_pdf", "statement_file"}


def _is_salary_credit_for_next_month(entry: Dict[str, object]) -> bool:
    tx_type = str(entry.get("type") or "").strip().lower()
    category = str(entry.get("category") or "").strip().lower()
    tx_date = entry.get("date")

    if tx_type != "income" or category != "salary":
        return False
    if not isinstance(tx_date, datetime):
        return False

    # Many payroll cycles credit in the last week for the following month.
    return 25 <= tx_date.day <= 31


def _parse_report_month(report_month: Optional[str]) -> tuple[datetime, datetime, str]:
    if report_month:
        try:
            month_start = datetime.strptime(report_month, "%Y-%m")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="report_month must be in YYYY-MM format") from exc
    else:
        now = datetime.utcnow()
        month_start = datetime(now.year, now.month, 1)

    last_day = calendar.monthrange(month_start.year, month_start.month)[1]
    month_end = datetime(month_start.year, month_start.month, last_day, 23, 59, 59)
    month_label = month_start.strftime("%B %Y")
    return month_start, month_end, month_label


def _format_inr(value: float) -> str:
    return f"Rs. {value:,.0f}"


def _generate_bar_chart_svg(title: str, items: List[tuple[str, float]], color: str) -> str:
    width = 640
    row_height = 34
    chart_height = max(140, 70 + (len(items) * row_height))
    max_value = max((value for _, value in items), default=1) or 1
    bars = []

    for index, (label, value) in enumerate(items[:6]):
        y = 42 + index * row_height
        bar_width = int((value / max_value) * 360) if max_value else 0
        safe_label = label[:24]
        bars.append(
            f"""
            <text x="20" y="{y + 12}" font-size="12" fill="#334155">{safe_label}</text>
            <rect x="180" y="{y}" width="360" height="12" rx="6" fill="#e2e8f0"></rect>
            <rect x="180" y="{y}" width="{bar_width}" height="12" rx="6" fill="{color}"></rect>
            <text x="550" y="{y + 12}" font-size="12" text-anchor="end" fill="#0f172a">{_format_inr(value)}</text>
            """
        )

    return f"""
    <div style="margin: 24px 0;">
        <div style="font-weight: 700; font-size: 16px; color: #0f172a; margin-bottom: 10px;">{title}</div>
        <svg width="{width}" height="{chart_height}" viewBox="0 0 {width} {chart_height}" xmlns="http://www.w3.org/2000/svg" style="max-width: 100%; background: #f8fafc; border-radius: 16px;">
            {''.join(bars) if bars else '<text x="20" y="44" font-size="13" fill="#64748b">No data available</text>'}
        </svg>
    </div>
    """


def _generate_line_chart_svg(title: str, items: List[tuple[str, float]]) -> str:
    width = 640
    height = 240
    padding_left = 40
    padding_right = 24
    padding_top = 24
    padding_bottom = 36
    max_value = max((value for _, value in items), default=1) or 1
    usable_width = width - padding_left - padding_right
    usable_height = height - padding_top - padding_bottom

    if len(items) <= 1:
        points = f"{padding_left},{padding_top + usable_height}"
    else:
        point_list = []
        for index, (_, value) in enumerate(items):
            x = padding_left + (usable_width * index / (len(items) - 1))
            y = padding_top + usable_height - ((value / max_value) * usable_height if max_value else 0)
            point_list.append(f"{x:.1f},{y:.1f}")
        points = " ".join(point_list)

    labels = []
    for index, (label, _) in enumerate(items[:8]):
        x = padding_left + (usable_width * index / max(1, len(items[:8]) - 1))
        labels.append(f'<text x="{x:.1f}" y="{height - 12}" font-size="11" text-anchor="middle" fill="#64748b">{label}</text>')

    return f"""
    <div style="margin: 24px 0;">
        <div style="font-weight: 700; font-size: 16px; color: #0f172a; margin-bottom: 10px;">{title}</div>
        <svg width="{width}" height="{height}" viewBox="0 0 {width} {height}" xmlns="http://www.w3.org/2000/svg" style="max-width: 100%; background: #f8fafc; border-radius: 16px;">
            <line x1="{padding_left}" y1="{padding_top + usable_height}" x2="{width - padding_right}" y2="{padding_top + usable_height}" stroke="#cbd5e1" stroke-width="1"/>
            <line x1="{padding_left}" y1="{padding_top}" x2="{padding_left}" y2="{padding_top + usable_height}" stroke="#cbd5e1" stroke-width="1"/>
            <polyline fill="none" stroke="#2563eb" stroke-width="3" points="{points}"/>
            {''.join(labels) if items else '<text x="20" y="44" font-size="13" fill="#64748b">No data available</text>'}
        </svg>
    </div>
    """


def _build_expense_report_html(user: User, month_label: str, expenses: List[Transaction]) -> str:
    total_expense = sum(float(tx.amount or 0) for tx in expenses)
    expense_count = len(expenses)
    avg_expense = total_expense / expense_count if expense_count else 0

    by_category = defaultdict(float)
    by_day = defaultdict(float)
    for tx in expenses:
        by_category[tx.category or "Other"] += float(tx.amount or 0)
        by_day[tx.date.strftime("%d %b")] += float(tx.amount or 0)

    category_items = sorted(by_category.items(), key=lambda item: item[1], reverse=True)
    day_items = sorted(by_day.items(), key=lambda item: datetime.strptime(item[0], "%d %b"))
    top_category = category_items[0][0] if category_items else "N/A"

    summary_cards = f"""
    <div style="display: flex; gap: 12px; flex-wrap: wrap; margin: 20px 0 10px;">
        <div style="flex: 1 1 180px; background: #eff6ff; border-radius: 16px; padding: 16px;">
            <div style="font-size: 12px; color: #64748b; text-transform: uppercase; letter-spacing: .08em;">Total Expense</div>
            <div style="font-size: 28px; font-weight: 700; color: #0f172a;">{_format_inr(total_expense)}</div>
        </div>
        <div style="flex: 1 1 180px; background: #fef2f2; border-radius: 16px; padding: 16px;">
            <div style="font-size: 12px; color: #64748b; text-transform: uppercase; letter-spacing: .08em;">Transactions</div>
            <div style="font-size: 28px; font-weight: 700; color: #0f172a;">{expense_count}</div>
        </div>
        <div style="flex: 1 1 180px; background: #ecfeff; border-radius: 16px; padding: 16px;">
            <div style="font-size: 12px; color: #64748b; text-transform: uppercase; letter-spacing: .08em;">Average Expense</div>
            <div style="font-size: 28px; font-weight: 700; color: #0f172a;">{_format_inr(avg_expense)}</div>
        </div>
        <div style="flex: 1 1 180px; background: #f8fafc; border-radius: 16px; padding: 16px;">
            <div style="font-size: 12px; color: #64748b; text-transform: uppercase; letter-spacing: .08em;">Top Category</div>
            <div style="font-size: 24px; font-weight: 700; color: #0f172a;">{top_category}</div>
        </div>
    </div>
    """

    category_rows = "".join(
        f"<tr><td style='padding: 8px 0; color: #334155;'>{label}</td><td style='padding: 8px 0; text-align: right; font-weight: 600; color: #0f172a;'>{_format_inr(value)}</td></tr>"
        for label, value in category_items[:8]
    ) or "<tr><td colspan='2' style='padding: 8px 0; color: #64748b;'>No expense data available.</td></tr>"

    return f"""
    <html>
    <body style="margin:0; padding:0; background:#f1f5f9; font-family:Arial,sans-serif; color:#0f172a;">
        <div style="max-width:760px; margin:0 auto; padding:24px;">
            <div style="background:linear-gradient(135deg, #1d4ed8 0%, #2563eb 100%); color:white; border-radius:24px; padding:28px;">
                <div style="font-size:13px; letter-spacing:.12em; text-transform:uppercase; opacity:.85;">Ledger App Expense Report</div>
                <h1 style="margin:10px 0 6px; font-size:32px;">{month_label}</h1>
                <div style="font-size:15px; opacity:.92;">Prepared for {user.full_name or user.username}</div>
            </div>
            <div style="background:white; border-radius:24px; padding:24px; margin-top:20px;">
                {summary_cards}
                {_generate_bar_chart_svg("Top Expense Categories", category_items[:6], "#ef4444")}
                {_generate_line_chart_svg("Daily Expense Trend", day_items[:8])}
                <div style="margin-top: 24px;">
                    <div style="font-weight:700; font-size:16px; color:#0f172a; margin-bottom:10px;">Category Summary</div>
                    <table style="width:100%; border-collapse:collapse;">
                        {category_rows}
                    </table>
                </div>
            </div>
        </div>
    </body>
    </html>
    """


def _send_report_email(recipient_email: str, subject: str, html_body: str) -> None:
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_username = os.getenv("SMTP_USERNAME")
    smtp_password = os.getenv("SMTP_PASSWORD")
    smtp_from = os.getenv("SMTP_FROM_EMAIL") or smtp_username

    if not smtp_username or not smtp_password or not smtp_from:
        raise HTTPException(
            status_code=500,
            detail="Email delivery is not configured. Set SMTP_USERNAME, SMTP_PASSWORD, and SMTP_FROM_EMAIL in backend/.env.",
        )

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = smtp_from
    message["To"] = recipient_email
    message.set_content("Your email client does not support HTML reports.")
    message.add_alternative(html_body, subtype="html")

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as smtp:
            smtp.starttls()
            smtp.login(smtp_username, smtp_password)
            smtp.send_message(message)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to send report email: {exc}") from exc


GOOGLE_AUTH_BASE = "https://accounts.google.com/o/oauth2/v2/auth"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
GOOGLE_USERINFO_URL = "https://www.googleapis.com/oauth2/v2/userinfo"
GMAIL_SEND_URL = "https://gmail.googleapis.com/gmail/v1/users/me/messages/send"
GMAIL_MESSAGES_URL = "https://gmail.googleapis.com/gmail/v1/users/me/messages"
GOOGLE_GMAIL_SCOPES = [
    "openid",
    "email",
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/gmail.readonly",
]

# Bank alert email category detection
_BANK_CATEGORY_KEYWORDS: dict = {
    "Food": ["swiggy", "zomato", "uber eat", "restaurant", "cafe", "domino", "pizza", "mcdonald", "kfc", "dining", "food court", "bakery"],
    "Shopping": ["amazon", "flipkart", "myntra", "ajio", "nykaa", "dmart", "reliance retail", "mall", "online shopping"],
    "Fuel": ["petrol", "diesel", "fuel", " bp ", "shell", " hp ", "indian oil", "bharat petroleum", "hpcl", "iocl"],
    "Bills & Utilities": ["electricity", "water bill", "gas bill", "bescom", "msedcl", "tata power", "bsnl", "airtel", "jio", "vi ", "vodafone", "idea", "internet", "broadband", "dth", "recharge"],
    "Healthcare": ["hospital", "pharmacy", "medical", "doctor", "clinic", "apollo", "fortis", "medplus", "1mg", "pharmeasy", "netmeds"],
    "Travel": ["uber", "ola ", "rapido", "irctc", "air india", "indigo", "spicejet", "flight", "train", "bus", "metro", "airport", "makemytrip", "goibibo"],
    "Entertainment": ["netflix", "prime video", "hotstar", "spotify", "youtube premium", "bookmyshow", "pvr", "inox", "movie"],
    "EMI / Loan": ["emi", "loan repayment", "mortgage", "installment"],
    "Investment": ["mutual fund", "sip", "zerodha", "groww", "stock", "equity", "nse", "bse", "nps"],
    "Salary": ["salary", "credited by employer", "payroll"],
    "Interest Income": ["interest credit", "interest income", "fd interest", "recurring deposit"],
    "Cashback / Refund": ["cashback", "refund", "reversal"],
}


def _detect_bank_alert_category(text: str) -> str:
    """Return the most likely expense/income category based on keywords."""
    lowered = text.lower()
    for category, keywords in _BANK_CATEGORY_KEYWORDS.items():
        if any(kw in lowered for kw in keywords):
            return category
    return "Other"


def _parse_bank_alert_email(subject: str, body: str) -> Optional[dict]:
    """Parse a bank alert email and return transaction fields, or None."""
    text = re.sub(r"\s+", " ", (subject + " " + body))

    # --- Detect debit vs credit ---
    debit_re = re.compile(
        r"\b(debit(?:ed)?|spent|paid|withdrawn|withdrawal|purchase|charged|debited from)\b",
        re.IGNORECASE,
    )
    credit_re = re.compile(
        r"\b(credit(?:ed)?|received|deposit(?:ed)?|refund(?:ed)?|cashback|salary|bonus|dividend)\b",
        re.IGNORECASE,
    )
    is_debit = bool(debit_re.search(text))
    is_credit = bool(credit_re.search(text))

    if not is_debit and not is_credit:
        return None

    # Salary/bonus/interest keywords always resolve to income
    if re.search(r"\b(salary|payroll|bonus|dividend|interest credit|refund)\b", text, re.IGNORECASE):
        tx_type = "income"
    elif is_debit:
        tx_type = "expense"
    else:
        tx_type = "income"

    # --- Extract amount (handle INR / Rs / ₹ / $ / USD / EUR) ---
    amount_patterns = [
        r"(?:INR|Rs\.?|\u20b9|USD|\$|EUR|\u20ac)\s*([0-9,]+(?:\.[0-9]{1,2})?)",
        r"([0-9,]+(?:\.[0-9]{1,2})?)\s*(?:INR|Rs\.?|\u20b9)",
    ]
    amount: Optional[float] = None
    for pat in amount_patterns:
        m = re.search(pat, text, re.IGNORECASE)
        if m:
            try:
                amount = float(m.group(1).replace(",", ""))
                break
            except ValueError:
                continue
    if not amount or amount <= 0:
        return None

    # --- Extract merchant / description ---
    desc_patterns = [
        r"(?:at|to|from|merchant:|for)\s+([A-Z][A-Za-z0-9\s\-&'\.]{2,40}?)(?=\s+(?:on|via|using|with|ref|txn|by|\.|,)|$)",
        r"(?:UPI-|VPA|trf to|transfer to)\s*([A-Za-z0-9@\.\-_]{3,40})",
        r"(?:purchase at|payment to|paid to)\s+([A-Za-z0-9\s\-&'\.]{2,40})",
    ]
    description = subject.strip()[:200]
    for pat in desc_patterns:
        m = re.search(pat, text)
        if m:
            desc = m.group(1).strip()
            if len(desc) >= 3:
                description = desc
                break
    description = re.sub(r"\s+", " ", description).strip()[:200]

    category = _detect_bank_alert_category(description + " " + text)
    return {"type": tx_type, "amount": amount, "description": description, "category": category}


def _gmail_list_messages(access_token: str, query: str, max_results: int = 100) -> list:
    """List Gmail message IDs matching query."""
    url = f"{GMAIL_MESSAGES_URL}?q={urlencode({'q': query})}&maxResults={max_results}"
    data = _google_api_json(url, headers={"Authorization": f"Bearer {access_token}"})
    return data.get("messages", [])


def _gmail_get_message(access_token: str, message_id: str) -> dict:
    """Fetch a single Gmail message (metadata + snippet)."""
    url = f"{GMAIL_MESSAGES_URL}/{message_id}?format=full"
    return _google_api_json(url, headers={"Authorization": f"Bearer {access_token}"})


def _decode_gmail_body(payload: dict) -> str:
    """Recursively decode the body of a Gmail message payload."""
    mime_type = payload.get("mimeType", "")
    body_data = (payload.get("body") or {}).get("data", "")

    if body_data:
        try:
            return base64.urlsafe_b64decode(body_data + "==").decode("utf-8", errors="ignore")
        except Exception:
            pass

    if "parts" in payload:
        # Prefer text/plain, fallback to text/html
        for preferred in ("text/plain", "text/html"):
            for part in payload["parts"]:
                if part.get("mimeType") == preferred:
                    text = _decode_gmail_body(part)
                    if text:
                        return text
        # Recurse into any part
        for part in payload["parts"]:
            text = _decode_gmail_body(part)
            if text:
                return text
    return ""


def _google_client_id() -> str:
    value = os.getenv("GOOGLE_CLIENT_ID")
    if not value:
        raise HTTPException(status_code=500, detail="Google OAuth is not configured. Set GOOGLE_CLIENT_ID in backend/.env.")
    return value


def _google_client_secret() -> str:
    value = os.getenv("GOOGLE_CLIENT_SECRET")
    if not value:
        raise HTTPException(status_code=500, detail="Google OAuth is not configured. Set GOOGLE_CLIENT_SECRET in backend/.env.")
    return value


def _google_redirect_uri(request: Optional[Request] = None) -> str:
    configured = os.getenv("GOOGLE_REDIRECT_URI")
    if configured:
        return configured
    if request is None:
        raise HTTPException(status_code=500, detail="Google OAuth redirect URI is not configured.")
    base = str(request.base_url).rstrip("/")
    return f"{base}/api/integrations/gmail/callback"


def _google_api_json(url: str, method: str = "GET", data: Optional[dict] = None, headers: Optional[dict] = None):
    encoded_data = None
    request_headers = headers.copy() if headers else {}
    if data is not None:
        encoded_data = json.dumps(data).encode("utf-8")
        request_headers.setdefault("Content-Type", "application/json")

    req = UrlRequest(url, data=encoded_data, headers=request_headers, method=method)
    try:
        with urlopen(req, timeout=20) as response:
            payload = response.read().decode("utf-8")
            return json.loads(payload) if payload else {}
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=500, detail=f"Google API error: {body or exc.reason}") from exc
    except URLError as exc:
        raise HTTPException(status_code=500, detail=f"Google API connection failed: {exc.reason}") from exc


def _google_form_post(url: str, payload: dict):
    data = urlencode(payload).encode("utf-8")
    req = UrlRequest(url, data=data, headers={"Content-Type": "application/x-www-form-urlencoded"}, method="POST")
    try:
        with urlopen(req, timeout=20) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=500, detail=f"Google OAuth error: {body or exc.reason}") from exc
    except URLError as exc:
        raise HTTPException(status_code=500, detail=f"Google OAuth connection failed: {exc.reason}") from exc


def _create_google_oauth_state(user_id: int) -> str:
    expires = datetime.utcnow() + timedelta(minutes=10)
    return jwt.encode(
        {"sub": str(user_id), "purpose": "gmail_oauth", "exp": expires, "iat": datetime.utcnow()},
        SECRET_KEY,
        algorithm=ALGORITHM,
    )


def _decode_google_oauth_state(state: str) -> int:
    try:
        payload = jwt.decode(state, SECRET_KEY, algorithms=[ALGORITHM])
    except JWTError as exc:
        raise HTTPException(status_code=400, detail="Invalid OAuth state.") from exc

    if payload.get("purpose") != "gmail_oauth":
        raise HTTPException(status_code=400, detail="Invalid OAuth state purpose.")
    return int(payload.get("sub"))


def _upsert_gmail_integration(db: Session, user_id: int, token_data: dict, account_email: str) -> Integration:
    integration = db.query(Integration).filter(
        Integration.user_id == user_id,
        Integration.app_name == "gmail"
    ).first()

    token_json = json.dumps(token_data)
    if integration:
        integration.connected = True
        integration.account_email = account_email
        integration.oauth_token = token_json
        integration.last_sync = datetime.utcnow()
    else:
        integration = Integration(
            user_id=user_id,
            app_name="gmail",
            connected=True,
            account_email=account_email,
            oauth_token=token_json,
            sync_frequency="manual",
            last_sync=datetime.utcnow(),
        )
        db.add(integration)

    db.commit()
    db.refresh(integration)
    return integration


def _get_gmail_integration(db: Session, user_id: int) -> Integration:
    integration = db.query(Integration).filter(
        Integration.user_id == user_id,
        Integration.app_name == "gmail",
        Integration.connected == True
    ).first()
    if not integration or not integration.oauth_token:
        raise HTTPException(status_code=400, detail="Gmail is not connected. Please connect your Gmail account first.")
    return integration


def _get_valid_google_access_token(db: Session, integration: Integration) -> str:
    token_data = json.loads(integration.oauth_token or "{}")
    access_token = token_data.get("access_token")
    refresh_token = token_data.get("refresh_token")
    expires_at = token_data.get("expires_at")

    is_expired = True
    if expires_at:
        try:
            is_expired = datetime.utcnow() >= datetime.fromisoformat(expires_at)
        except ValueError:
            is_expired = True

    if access_token and not is_expired:
        return access_token

    if not refresh_token:
        raise HTTPException(status_code=400, detail="Gmail connection has expired. Please reconnect your Gmail account.")

    refreshed = _google_form_post(GOOGLE_TOKEN_URL, {
        "client_id": _google_client_id(),
        "client_secret": _google_client_secret(),
        "refresh_token": refresh_token,
        "grant_type": "refresh_token",
    })

    token_data["access_token"] = refreshed["access_token"]
    token_data["expires_at"] = (datetime.utcnow() + timedelta(seconds=int(refreshed.get("expires_in", 3600)) - 60)).isoformat()
    integration.oauth_token = json.dumps(token_data)
    integration.last_sync = datetime.utcnow()
    db.commit()
    db.refresh(integration)
    return token_data["access_token"]


def _send_gmail_message(db: Session, integration: Integration, recipient_email: str, subject: str, html_body: str) -> None:
    access_token = _get_valid_google_access_token(db, integration)
    message = EmailMessage()
    message["To"] = recipient_email
    message["From"] = integration.account_email or "me"
    message["Subject"] = subject
    message.set_content("Your email client does not support HTML reports.")
    message.add_alternative(html_body, subtype="html")

    raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode("utf-8")
    _google_api_json(
        GMAIL_SEND_URL,
        method="POST",
        data={"raw": raw_message},
        headers={"Authorization": f"Bearer {access_token}"},
    )


EXPORT_COLUMNS = [
    "section", "id", "name", "type", "description", "value", "amount", "current_value",
    "category", "date", "notes", "recurring", "annual_growth_rate", "spread_over_year", "source", "lender", "outstanding",
    "is_loan", "loan_start_date", "loan_tenure_months", "interest_rate", "opportunity_cost_rate", "monthly_payment", "linked_asset_id", "due_date", "include_in_balance",
    "include_in_income", "loan_emi_linked", "limit", "period", "start_month", "target", "current", "target_date", "metric", "metric_value"
]


def _bool_to_csv(value) -> str:
    return "true" if bool(value) else "false"


def _csv_to_bool(value, default=False) -> bool:
    if value is None or value == "":
        return default
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _csv_to_float(value, default=0.0):
    if value is None or value == "":
        return default
    return float(value)


def _csv_to_datetime(value):
    if not value:
        return None
    return datetime.fromisoformat(value) if "T" in value else datetime.strptime(value, "%Y-%m-%d")


def _build_export_rows(current_user: User, db: Session):
    transactions = db.query(Transaction).filter(Transaction.user_id == current_user.id).order_by(Transaction.date.asc()).all()
    investments = db.query(Investment).filter(Investment.user_id == current_user.id).order_by(Investment.created_at.asc()).all()
    liabilities = db.query(Liability).filter(Liability.user_id == current_user.id).order_by(Liability.created_at.asc()).all()
    assets = db.query(Asset).filter(Asset.user_id == current_user.id).order_by(Asset.created_at.asc()).all()
    budgets = db.query(Budget).filter(Budget.user_id == current_user.id).order_by(Budget.created_at.asc()).all()
    goals = db.query(Goal).filter(Goal.user_id == current_user.id).order_by(Goal.created_at.asc()).all()

    income_total = sum(float(t.amount or 0) for t in transactions if t.type == TransactionType.INCOME)
    expense_total = sum(float(t.amount or 0) for t in transactions if t.type == TransactionType.EXPENSE)
    investment_total = sum(float(i.amount_invested or 0) for i in investments)
    asset_total = sum(float(a.value or 0) for a in assets)
    liability_total = sum(float(l.outstanding or l.amount or 0) for l in liabilities)
    savings_total = income_total + investment_total - expense_total

    rows = []

    for tx in transactions:
        rows.append({
            "section": "transaction",
            "id": tx.id,
            "type": tx.type.value if hasattr(tx.type, "value") else str(tx.type),
            "description": tx.description,
            "amount": tx.amount,
            "category": tx.category,
            "date": tx.date.isoformat() if tx.date else "",
            "notes": tx.notes or "",
            "recurring": _bool_to_csv(tx.recurring),
            "spread_over_year": _bool_to_csv(getattr(tx, "spread_over_year", False)),
            "source": tx.source or "",
        })

    for inv in investments:
        rows.append({
            "section": "investment",
            "id": inv.id,
            "name": inv.name,
            "type": inv.type,
            "amount": inv.amount_invested,
            "current_value": inv.current_value if inv.current_value is not None else "",
            "annual_growth_rate": inv.annual_growth_rate if getattr(inv, "annual_growth_rate", None) is not None else "",
            "notes": inv.notes or "",
            "recurring": _bool_to_csv(getattr(inv, "monthly_sip", False)),
        })

    for lib in liabilities:
        rows.append({
            "section": "liability",
            "id": lib.id,
            "lender": lib.lender,
            "amount": lib.amount,
            "outstanding": lib.outstanding,
            "is_loan": _bool_to_csv(getattr(lib, "is_loan", False)),
            "loan_start_date": lib.loan_start_date.isoformat() if getattr(lib, "loan_start_date", None) else "",
            "loan_tenure_months": getattr(lib, "loan_tenure_months", None) if getattr(lib, "loan_tenure_months", None) is not None else "",
            "interest_rate": lib.interest_rate if lib.interest_rate is not None else "",
            "opportunity_cost_rate": getattr(lib, "opportunity_cost_rate", None) if getattr(lib, "opportunity_cost_rate", None) is not None else "",
            "monthly_payment": lib.monthly_payment if lib.monthly_payment is not None else "",
            "linked_asset_id": getattr(lib, "linked_asset_id", None) if getattr(lib, "linked_asset_id", None) is not None else "",
            "due_date": lib.due_date.isoformat() if lib.due_date else "",
            "notes": lib.notes or "",
        })

    for asset in assets:
        rows.append({
            "section": "asset",
            "id": asset.id,
            "name": asset.name,
            "type": asset.type or "",
            "value": asset.value,
            "description": asset.description or "",
            "include_in_balance": _bool_to_csv(asset.include_in_balance),
            "include_in_income": _bool_to_csv(getattr(asset, "include_in_income", False)),
            "loan_emi_linked": _bool_to_csv(getattr(asset, "loan_emi_linked", False)),
        })

    for budget in budgets:
        rows.append({
            "section": "budget",
            "id": budget.id,
            "category": budget.category,
            "limit": budget.limit,
            "period": getattr(budget, 'period', 'monthly') or 'monthly',
            "recurring": _bool_to_csv(getattr(budget, 'recurring', False)),
            "start_month": getattr(budget, 'start_month', None) or "",
        })

    for goal in goals:
        rows.append({
            "section": "goal",
            "id": goal.id,
            "name": goal.name,
            "target": goal.target,
            "current": goal.current,
            "target_date": goal.target_date.isoformat() if goal.target_date else "",
        })

    summary_rows = [
        ("income_total", income_total),
        ("expense_total", expense_total),
        ("investment_total", investment_total),
        ("asset_total", asset_total),
        ("liability_total", liability_total),
        ("savings_total", savings_total),
    ]
    for metric, metric_value in summary_rows:
        rows.append({
            "section": "summary",
            "metric": metric,
            "metric_value": metric_value,
        })

    return rows


# ==================== Authentication Endpoints ====================

@app.post("/api/auth/register", response_model=UserSchema)
async def register(user: UserCreate, request: Request, db: Session = Depends(get_db)):
    """Register a new user"""
    # Get client IP
    client_ip = get_client_ip(request)
    
    # Check rate limit
    if not check_registration_rate_limit(client_ip):
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many registration attempts. Please try again later."
        )
    
    # Check if user exists
    db_user = db.query(User).filter(
        (User.email == user.email) | (User.username == user.username)
    ).first()
    if db_user:
        raise HTTPException(status_code=400, detail="Email or username already registered")
    
    # Create new user
    hashed_password = get_password_hash(user.password)
    user_count = db.query(User).count()
    db_user = User(
        email=user.email,
        username=user.username,
        full_name=user.full_name,
        hashed_password=hashed_password,
        role=UserRole.ADMIN.value if user_count == 0 else UserRole.USER.value,
        permissions_json=json.dumps(get_default_permissions(UserRole.ADMIN.value if user_count == 0 else UserRole.USER.value))
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    
    # Log audit
    log_audit(db, db_user.id, "create", "user", db_user.id, 
              f"User registered: {user.username}")
    
    return serialize_user(db_user)


@app.post("/api/auth/login", response_model=Token)
async def login(
    request: Request,
    db: Session = Depends(get_db)
):
    """Login and get access token"""
    # Get client IP
    client_ip = get_client_ip(request)

    # Support both application/x-www-form-urlencoded and application/json credentials
    username = None
    password = None

    content_type = (request.headers.get("content-type") or "").split(";")[0].strip().lower()
    if content_type == "application/json":
        payload = await request.json()
        username = payload.get("username")
        password = payload.get("password")
    else:
        form_data = await request.form()
        username = form_data.get("username")
        password = form_data.get("password")

    if not username or not password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Username and password are required"
        )

    # Authenticate user
    user = authenticate_user(db, username, password, client_ip)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/api/auth/me", response_model=UserSchema)
def read_users_me(current_user: User = Depends(get_current_active_user)):
    """Get current user"""
    return serialize_user(current_user)


@app.put("/api/auth/me/profile", response_model=UserSchema)
def update_my_profile(
    profile_update: UserProfileUpdate,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Allow any authenticated user to edit their own profile."""
    changed_fields = {}

    if profile_update.email is not None:
        normalized_email = profile_update.email.strip().lower()
        if not normalized_email:
            raise HTTPException(status_code=400, detail="Email cannot be empty")

        existing = db.query(User).filter(User.email == normalized_email, User.id != current_user.id).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email is already used by another account")

        if current_user.email != normalized_email:
            changed_fields["email"] = {"from": current_user.email, "to": normalized_email}
            current_user.email = normalized_email

    if profile_update.full_name is not None:
        normalized_name = profile_update.full_name.strip() or None
        if current_user.full_name != normalized_name:
            changed_fields["full_name"] = {"from": current_user.full_name, "to": normalized_name}
            current_user.full_name = normalized_name

    if not changed_fields:
        return serialize_user(current_user)

    db.commit()
    db.refresh(current_user)
    log_audit(
        db,
        current_user.id,
        "update",
        "user_profile",
        str(current_user.id),
        f"Updated profile for {current_user.username}",
        {"fields": changed_fields}
    )
    return serialize_user(current_user)


@app.post("/api/auth/me/password")
def change_my_password(
    password_change: UserPasswordChange,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Allow any authenticated user to change their own password."""
    current_password = (password_change.current_password or "").strip()
    new_password = (password_change.new_password or "").strip()

    if not current_password or not new_password:
        raise HTTPException(status_code=400, detail="Current and new password are required")
    if len(new_password) < 8:
        raise HTTPException(status_code=400, detail="New password must be at least 8 characters")
    if current_password == new_password:
        raise HTTPException(status_code=400, detail="New password must be different from current password")
    if not verify_password(current_password, current_user.hashed_password):
        raise HTTPException(status_code=403, detail="Current password is incorrect")

    current_user.hashed_password = get_password_hash(new_password)
    db.commit()

    log_audit(
        db,
        current_user.id,
        "update",
        "user_password",
        str(current_user.id),
        f"Changed password for {current_user.username}"
    )
    return {"message": "Password updated successfully"}


@app.post("/api/rbac/external/provision-user", response_model=UserSchema)
def external_provision_user_access(
    payload: ExternalRBACProvisionRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(default=None),
    x_rbac_api_key: Optional[str] = Header(default=None)
):
    """Provision or update user access from an external identity/authentication system."""
    verify_external_rbac_api_key(authorization, x_rbac_api_key)

    identity_provider = (payload.identity_provider or "").strip() or None
    external_subject = (payload.external_subject or "").strip() or None
    username = (payload.username or "").strip()
    email = (str(payload.email).strip().lower() if payload.email else "")

    db_user = None
    if external_subject and identity_provider:
        db_user = db.query(User).filter(
            User.external_subject == external_subject,
            User.identity_provider == identity_provider
        ).first()
    if not db_user and username:
        db_user = db.query(User).filter(User.username == username).first()
    if not db_user and email:
        db_user = db.query(User).filter(User.email == email).first()

    conflict_user_id = db_user.id if db_user else -1
    if username:
        username_conflict = db.query(User).filter(User.username == username, User.id != conflict_user_id).first()
        if username_conflict:
            raise HTTPException(status_code=400, detail="Username is already used by another account")
    if email:
        email_conflict = db.query(User).filter(User.email == email, User.id != conflict_user_id).first()
        if email_conflict:
            raise HTTPException(status_code=400, detail="Email is already used by another account")

    requested_role = payload.role.value
    permissions_obj = payload.permissions.model_dump() if payload.permissions else get_default_permissions(requested_role)

    if not db_user:
        if not payload.create_if_missing:
            raise HTTPException(status_code=404, detail="User not found and create_if_missing=false")
        if not username or not email:
            raise HTTPException(status_code=400, detail="username and email are required to create a user")

        db_user = User(
            username=username,
            email=email,
            full_name=payload.full_name,
            hashed_password=get_password_hash(os.getenv("FEDERATED_PLACEHOLDER_PASSWORD", "change-me-federated-user")),
            role=requested_role,
            permissions_json=json.dumps(permissions_obj),
            is_active=payload.is_active,
            identity_provider=identity_provider,
            external_subject=external_subject,
        )
        db.add(db_user)
        db.commit()
        db.refresh(db_user)
        log_audit(
            db,
            db_user.id,
            "create",
            "external_rbac_user",
            str(db_user.id),
            f"Provisioned user {db_user.username} via external RBAC API",
            {
                "identity_provider": identity_provider,
                "external_subject": external_subject,
                "source_ip": get_client_ip(request),
            }
        )
        return serialize_user(db_user)

    db_user.role = requested_role
    db_user.permissions_json = json.dumps(permissions_obj)
    db_user.is_active = payload.is_active
    if payload.full_name is not None:
        db_user.full_name = payload.full_name
    if email:
        db_user.email = email
    if username:
        db_user.username = username
    if identity_provider is not None:
        db_user.identity_provider = identity_provider
    if external_subject is not None:
        db_user.external_subject = external_subject

    db.commit()
    db.refresh(db_user)
    log_audit(
        db,
        db_user.id,
        "update",
        "external_rbac_user",
        str(db_user.id),
        f"Updated user {db_user.username} via external RBAC API",
        {
            "identity_provider": db_user.identity_provider,
            "external_subject": db_user.external_subject,
            "source_ip": get_client_ip(request),
        }
    )
    return serialize_user(db_user)


@app.post("/api/rbac/external/federated-sync", response_model=UserSchema)
def external_federated_claim_sync(
    payload: FederatedClaimSyncRequest,
    request: Request,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(default=None),
    x_rbac_api_key: Optional[str] = Header(default=None)
):
    """Resolve OAuth/federated claims into RBAC and upsert user access configuration."""
    verify_external_rbac_api_key(authorization, x_rbac_api_key)

    resolved_role = resolve_federated_role(payload.groups, payload.role_hint)
    resolved_permissions = payload.permissions_override.model_dump() if payload.permissions_override else get_default_permissions(resolved_role)

    provision_payload = ExternalRBACProvisionRequest(
        username=payload.username,
        email=payload.email,
        full_name=payload.full_name,
        role=UserRole(resolved_role),
        is_active=True,
        permissions=payload.permissions_override,
        create_if_missing=payload.create_if_missing,
        identity_provider=payload.identity_provider,
        external_subject=payload.external_subject,
    )

    user = external_provision_user_access(
        payload=provision_payload,
        request=request,
        db=db,
        authorization=authorization,
        x_rbac_api_key=x_rbac_api_key,
    )

    db_user = db.query(User).filter(User.id == user["id"]).first()
    if db_user:
        db_user.permissions_json = json.dumps(resolved_permissions)
        db.commit()
        db.refresh(db_user)
        log_audit(
            db,
            db_user.id,
            "update",
            "federated_rbac_sync",
            str(db_user.id),
            f"Applied federated RBAC mapping for {db_user.username}",
            {
                "groups": payload.groups,
                "role_hint": payload.role_hint.value if payload.role_hint else None,
                "claims": payload.claims,
                "resolved_role": resolved_role,
                "source_ip": get_client_ip(request),
            }
        )
        return serialize_user(db_user)

    raise HTTPException(status_code=500, detail="Federated sync succeeded but user could not be loaded")


@app.get("/api/users", response_model=List[UserSchema])
def list_users(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """List all users. Admin only."""
    if not user_has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="You do not have permission to manage users")
    return [serialize_user(user) for user in db.query(User).order_by(User.created_at.asc()).all()]


@app.put("/api/users/{user_id}/role", response_model=UserSchema)
def update_user_role(
    user_id: int,
    role_update: UserRoleUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Update a user's role. Admin only."""
    if not user_has_permission(current_user, "manage_roles") and current_user.role != UserRole.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail="You do not have permission to manage roles")
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    if db_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot change your own role")

    is_current_superadmin = current_user.role == UserRole.SUPERADMIN.value
    requested_role = role_update.role.value

    if requested_role == UserRole.SUPERADMIN.value and not is_current_superadmin:
        raise HTTPException(status_code=403, detail="Only the super admin can assign the super admin role")

    if db_user.role == UserRole.SUPERADMIN.value and not is_current_superadmin:
        raise HTTPException(status_code=403, detail="Only the super admin can modify the super admin account")

    if requested_role == UserRole.ADMIN.value and not is_current_superadmin and current_user.role != UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Only admins can assign the admin role")

    if current_user.role == UserRole.ADMIN.value and requested_role == UserRole.ADMIN.value:
        raise HTTPException(status_code=403, detail="Only the super admin can promote another user to admin")

    if db_user.role in {UserRole.ADMIN.value, UserRole.SUPERADMIN.value} and requested_role not in {UserRole.ADMIN.value, UserRole.SUPERADMIN.value}:
        protected_admin_count = db.query(User).filter(User.role.in_([UserRole.ADMIN.value, UserRole.SUPERADMIN.value]), User.is_active == True).count()
        if protected_admin_count <= 1:
            raise HTTPException(status_code=400, detail="At least one active admin-level account must remain")

    db_user.role = requested_role
    db_user.permissions_json = json.dumps(get_default_permissions(requested_role))
    db.commit()
    db.refresh(db_user)
    log_audit(db, current_user.id, "update", "user_role", str(db_user.id), f"Updated role for {db_user.username} to {requested_role}")
    return serialize_user(db_user)


@app.put("/api/users/{user_id}/status", response_model=UserSchema)
def update_user_status(
    user_id: int,
    status_update: UserStatusUpdate,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Block or unblock a user. Admin only."""
    if not user_has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="You do not have permission to manage users")
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    if db_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot block your own account")
    if db_user.role == UserRole.SUPERADMIN.value and current_user.role != UserRole.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail="Only the super admin can block the super admin account")
    if db_user.role in {UserRole.ADMIN.value, UserRole.SUPERADMIN.value} and status_update.is_active is False:
        admin_count = db.query(User).filter(User.role.in_([UserRole.ADMIN.value, UserRole.SUPERADMIN.value]), User.is_active == True).count()
        if admin_count <= 1:
            raise HTTPException(status_code=400, detail="At least one active admin-level account must remain")

    db_user.is_active = status_update.is_active
    db.commit()
    db.refresh(db_user)
    log_audit(db, current_user.id, "update", "user_status", str(db_user.id), f"Updated status for {db_user.username} to {'active' if status_update.is_active else 'blocked'}")
    return serialize_user(db_user)


@app.put("/api/users/{user_id}/permissions", response_model=UserSchema)
def update_user_permissions(
    user_id: int,
    permissions_update: UserPermissionsUpdate,
    current_user: User = Depends(require_superadmin),
    db: Session = Depends(get_db)
):
    """Update a user's tab/page/field/permission access. Super admin only."""
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")

    db_user.permissions_json = json.dumps(permissions_update.permissions.model_dump())
    db.commit()
    db.refresh(db_user)
    log_audit(db, current_user.id, "update", "user_permissions", str(db_user.id), f"Updated permissions for {db_user.username}")
    return serialize_user(db_user)


@app.delete("/api/users/{user_id}")
def delete_user(
    user_id: int,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Delete a user account. Admin only."""
    if not user_has_permission(current_user, "manage_users"):
        raise HTTPException(status_code=403, detail="You do not have permission to manage users")
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    if db_user.id == current_user.id:
        raise HTTPException(status_code=400, detail="You cannot delete your own account")
    if db_user.role == UserRole.SUPERADMIN.value and current_user.role != UserRole.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail="Only the super admin can delete the super admin account")
    if db_user.role in {UserRole.ADMIN.value, UserRole.SUPERADMIN.value}:
        admin_count = db.query(User).filter(User.role.in_([UserRole.ADMIN.value, UserRole.SUPERADMIN.value]), User.is_active == True).count()
        if admin_count <= 1:
            raise HTTPException(status_code=400, detail="At least one active admin-level account must remain")

    db.delete(db_user)
    db.commit()
    return {"message": "User deleted successfully"}


@app.post("/api/admin/reset-all-data", response_model=AdminResetResponse)
def reset_all_financial_data(
    payload: AdminResetRequest,
    request: Request,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Reset all financial data across the system. Admin password confirmation required."""
    if not user_has_permission(current_user, "reset_data"):
        raise HTTPException(status_code=403, detail="You do not have permission to reset financial data")

    if not payload.password or not verify_password(payload.password, current_user.hashed_password):
        raise HTTPException(status_code=403, detail="Incorrect admin password")

    document_rows = db.query(Document).all()
    files_deleted = 0
    for document in document_rows:
        file_path = document.file_path
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
                files_deleted += 1
            except OSError:
                pass

    counts = {
        "transactions_deleted": db.query(Transaction).count(),
        "budgets_deleted": db.query(Budget).count(),
        "goals_deleted": db.query(Goal).count(),
        "investments_deleted": db.query(Investment).count(),
        "liabilities_deleted": db.query(Liability).count(),
        "assets_deleted": db.query(Asset).count(),
        "documents_deleted": len(document_rows),
        "integrations_deleted": db.query(Integration).count(),
        "files_deleted": files_deleted,
    }

    try:
        db.query(Transaction).delete(synchronize_session=False)
        db.query(Budget).delete(synchronize_session=False)
        db.query(Goal).delete(synchronize_session=False)
        db.query(Investment).delete(synchronize_session=False)
        db.query(Liability).delete(synchronize_session=False)
        db.query(Asset).delete(synchronize_session=False)
        db.query(Document).delete(synchronize_session=False)
        db.query(Integration).delete(synchronize_session=False)
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Reset failed: {type(exc).__name__}: {exc}")

    log_audit(
        db,
        current_user.id,
        "delete",
        "system_financial_data",
        "global-reset",
        "Admin reset all financial data",
        details=counts,
        user_agent=request.headers.get("user-agent")
    )

    return AdminResetResponse(
        message="All financial data has been reset successfully.",
        **counts,
    )


@app.get("/api/admin/app-insights")
def get_admin_app_insights(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Admin-only runtime insights for frontend, backend and DB instances."""
    if not user_has_permission(current_user, "view_app_insights"):
        raise HTTPException(status_code=403, detail="You do not have permission to view app insights")

    # Backend instance metrics
    usage = resource.getrusage(resource.RUSAGE_SELF)
    # Linux ru_maxrss is in KB; on macOS it is bytes. Use a conservative conversion.
    backend_memory_mb = float(usage.ru_maxrss) / 1024.0
    if backend_memory_mb > 1024 * 1024:  # likely bytes input
        backend_memory_mb = float(usage.ru_maxrss) / (1024.0 * 1024.0)

    latency_samples = max(1, APP_INSIGHTS_STATE["backend_latency_samples"])
    avg_latency_ms = APP_INSIGHTS_STATE["backend_latency_ms_total"] / latency_samples

    pool_size = None
    pool_checked_out = None
    try:
        if hasattr(engine.pool, "size"):
            pool_size = int(engine.pool.size())
        if hasattr(engine.pool, "checkedout"):
            pool_checked_out = int(engine.pool.checkedout())
    except Exception:
        pool_size = None
        pool_checked_out = None

    backend_instance = {
        "status": "ok",
        "cpu_percent": None,  # Requires host-level telemetry tooling.
        "memory_mb": round(backend_memory_mb, 2),
        "connection_count": APP_INSIGHTS_STATE["active_requests"],
        "errors": APP_INSIGHTS_STATE["errors_5xx"],
        "discards": APP_INSIGHTS_STATE["discards_4xx"],
        "details": {
            "requests_total": APP_INSIGHTS_STATE["total_requests"],
            "avg_latency_ms": round(avg_latency_ms, 2),
            "db_pool_size": pool_size,
            "db_pool_checked_out": pool_checked_out,
            "uptime_seconds": int((datetime.utcnow() - APP_INSIGHTS_STATE["started_at"]).total_seconds()),
            "recent_errors": APP_INSIGHTS_STATE["last_errors"][-10:],
        },
    }

    # Frontend instance metrics (best-effort via nginx stub_status)
    frontend_instance = _fetch_frontend_nginx_status()
    frontend_instance.setdefault("cpu_percent", None)
    frontend_instance.setdefault("memory_mb", None)

    # DB instance metrics (MySQL global status)
    mysql_vars = [
        "Threads_connected",
        "Threads_running",
        "Aborted_clients",
        "Aborted_connects",
        "Connection_errors_max_connections",
        "Questions",
        "Uptime",
        "Bytes_received",
        "Bytes_sent",
        "Slow_queries",
        "Select_full_join",
        "Created_tmp_disk_tables",
    ]
    mysql_stats = _get_mysql_status_map(db, mysql_vars)
    db_errors = (
        mysql_stats.get("Aborted_clients", 0)
        + mysql_stats.get("Aborted_connects", 0)
        + mysql_stats.get("Connection_errors_max_connections", 0)
    )
    db_instance = {
        "status": "ok",
        "cpu_percent": None,  # Requires node/container runtime access.
        "memory_mb": None,    # Requires node/container runtime access.
        "connection_count": mysql_stats.get("Threads_connected", 0),
        "errors": db_errors,
        "discards": mysql_stats.get("Aborted_clients", 0),
        "details": {
            "threads_running": mysql_stats.get("Threads_running", 0),
            "questions_total": mysql_stats.get("Questions", 0),
            "uptime_seconds": mysql_stats.get("Uptime", 0),
            "aborted_connects": mysql_stats.get("Aborted_connects", 0),
            "connection_errors_max_connections": mysql_stats.get("Connection_errors_max_connections", 0),
            "bytes_received": mysql_stats.get("Bytes_received", 0),
            "bytes_sent": mysql_stats.get("Bytes_sent", 0),
            "slow_queries": mysql_stats.get("Slow_queries", 0),
            "select_full_join": mysql_stats.get("Select_full_join", 0),
            "tmp_disk_tables": mysql_stats.get("Created_tmp_disk_tables", 0),
        },
    }

    alerts: List[Dict[str, str]] = []

    if backend_instance["errors"] > 0:
        alerts.append({
            "severity": "high",
            "instance": "backend",
            "message": f"Backend has {backend_instance['errors']} server error(s) since startup.",
        })
    if backend_instance["discards"] > 20:
        alerts.append({
            "severity": "medium",
            "instance": "backend",
            "message": f"Backend 4xx/discard count is elevated ({backend_instance['discards']}).",
        })
    if db_instance["connection_count"] > 80:
        alerts.append({
            "severity": "medium",
            "instance": "db",
            "message": f"DB active connections are high ({db_instance['connection_count']}).",
        })
    if db_instance["errors"] > 0:
        alerts.append({
            "severity": "high",
            "instance": "db",
            "message": f"DB reports connection/client abort errors ({db_instance['errors']}).",
        })
    if frontend_instance.get("status") != "ok":
        alerts.append({
            "severity": "medium",
            "instance": "frontend",
            "message": "Frontend status metrics are currently unavailable (nginx_status unreachable).",
        })
    elif frontend_instance.get("connection_count") and frontend_instance["connection_count"] > 300:
        alerts.append({
            "severity": "medium",
            "instance": "frontend",
            "message": f"Frontend active connections are high ({frontend_instance['connection_count']}).",
        })

    # Record snapshot for time-series history
    snapshot = {
        "ts": datetime.utcnow().isoformat(),
        "backend_connections": APP_INSIGHTS_STATE["active_requests"],
        "backend_errors": APP_INSIGHTS_STATE["errors_5xx"],
        "backend_discards": APP_INSIGHTS_STATE["discards_4xx"],
        "backend_latency_ms": round(avg_latency_ms, 2),
        "backend_requests_total": APP_INSIGHTS_STATE["total_requests"],
        "backend_memory_mb": round(backend_memory_mb, 2),
        "frontend_connections": frontend_instance.get("connection_count"),
        "frontend_errors": frontend_instance.get("errors"),
        "frontend_discards": frontend_instance.get("discards"),
        "frontend_reading": frontend_instance.get("reading"),
        "frontend_writing": frontend_instance.get("writing"),
        "frontend_waiting": frontend_instance.get("waiting"),
        "db_connections": mysql_stats.get("Threads_connected", 0),
        "db_errors": db_errors,
        "db_discards": mysql_stats.get("Aborted_clients", 0),
        "db_threads_running": mysql_stats.get("Threads_running", 0),
        "db_slow_queries": mysql_stats.get("Slow_queries", 0),
        "db_bytes_received": mysql_stats.get("Bytes_received", 0),
        "db_bytes_sent": mysql_stats.get("Bytes_sent", 0),
    }
    APP_INSIGHTS_HISTORY.append(snapshot)
    if len(APP_INSIGHTS_HISTORY) > MAX_INSIGHTS_HISTORY:
        del APP_INSIGHTS_HISTORY[:-MAX_INSIGHTS_HISTORY]

    instance_logs = _derive_instance_logs(
        history=list(APP_INSIGHTS_HISTORY),
        frontend_instance=frontend_instance,
        backend_instance=backend_instance,
        db_instance=db_instance,
        alerts=alerts,
        backend_recent_errors=backend_instance.get("details", {}).get("recent_errors", []),
    )

    return {
        "generated_at": datetime.utcnow(),
        "instances": {
            "frontend": frontend_instance,
            "backend": backend_instance,
            "db": db_instance,
        },
        "alerts": alerts,
        "history": list(APP_INSIGHTS_HISTORY),
        "instance_logs": instance_logs,
        "notes": [
            "Frontend CPU/memory and DB CPU/memory require host/container runtime integration and are reported as null for now.",
            "Frontend connection metrics come from nginx stub_status when available.",
        ],
    }


@app.get("/api/admin/app-insights/history")
def get_admin_app_insights_history(
    hours: Optional[float] = None,
    days: Optional[float] = None,
    weeks: Optional[float] = None,
    months: Optional[float] = None,
    start_time: Optional[str] = None,
    end_time: Optional[str] = None,
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Admin-only historical insights data with flexible time range filtering.
    
    Query Parameters:
    - hours: float (e.g., 0.5 for 30min, 24 for 24hrs)
    - days: float (e.g., 7 for 7 days)
    - weeks: float (e.g., 1 for 1 week)
    - months: float (e.g., 1 for 1 month)
    - start_time: ISO format datetime string (e.g., "2026-03-20T10:30:00")
    - end_time: ISO format datetime string (e.g., "2026-03-20T12:30:00")
    
    If no range is specified, returns last 24 hours by default.
    """
    if not user_has_permission(current_user, "view_app_insights"):
        raise HTTPException(status_code=403, detail="You do not have permission to view app insights")
    
    # Determine time range
    now = datetime.utcnow()
    
    if start_time and end_time:
        try:
            start = datetime.fromisoformat(start_time.replace("Z", "+00:00"))
            end = datetime.fromisoformat(end_time.replace("Z", "+00:00"))
        except (ValueError, AttributeError):
            raise HTTPException(status_code=400, detail="Invalid start_time or end_time format")
    else:
        # Calculate from relative time units
        total_seconds = 0
        if months:
            total_seconds += months * 30 * 24 * 3600  # Approximate
        if weeks:
            total_seconds += weeks * 7 * 24 * 3600
        if days:
            total_seconds += days * 24 * 3600
        if hours:
            total_seconds += hours * 3600
        
        if total_seconds <= 0:
            total_seconds = 24 * 3600  # Default to 24 hours
        
        start = now - timedelta(seconds=total_seconds)
        end = now
    
    # Query metrics from database
    metrics = db.query(AppInsightsMetric).filter(
        AppInsightsMetric.recorded_at >= start,
        AppInsightsMetric.recorded_at <= end
    ).order_by(AppInsightsMetric.recorded_at.asc()).all()
    
    if not metrics:
        return {
            "start_time": start.isoformat(),
            "end_time": end.isoformat(),
            "record_count": 0,
            "history": [],
            "instance_logs": {
                "frontend": [],
                "backend": [],
                "db": [],
            },
            "instances": {
                "frontend": {"status": "unknown", "connection_count": 0},
                "backend": {"status": "unknown", "connection_count": 0, "errors": 0, "discards": 0},
                "db": {"status": "unknown", "connection_count": 0, "errors": 0, "discards": 0},
            },
        }
    
    # Transform database records into the same format as the current snapshot
    history = []
    for metric in metrics:
        history.append({
            "ts": metric.recorded_at.isoformat(),
            "backend_connections": metric.backend_active_requests,
            "backend_errors": metric.backend_errors_5xx,
            "backend_discards": metric.backend_discards_4xx,
            "backend_latency_ms": metric.backend_avg_latency_ms,
            "backend_requests_total": metric.backend_requests_total,
            "backend_memory_mb": metric.backend_memory_mb,
            "frontend_connections": metric.frontend_active_connections,
            "frontend_errors": 0,
            "frontend_discards": 0,
            "frontend_reading": metric.frontend_reading,
            "frontend_writing": metric.frontend_writing,
            "frontend_waiting": metric.frontend_waiting,
            "db_connections": metric.db_connections,
            "db_errors": metric.db_errors,
            "db_discards": 0,
            "db_threads_running": metric.db_threads_running,
            "db_slow_queries": 0,
            "db_bytes_received": 0,
            "db_bytes_sent": 0,
        })
    
    # Calculate instances from the latest metric in the range
    latest_metric = metrics[-1]
    
    frontend_from_nginx = _fetch_frontend_nginx_status()
    frontend_instance = {
        "status": frontend_from_nginx.get("status", "ok"),
        "cpu_percent": None,
        "memory_mb": None,
        "connection_count": frontend_from_nginx.get("connection_count", latest_metric.frontend_active_connections),
        "reading": latest_metric.frontend_reading,
        "writing": latest_metric.frontend_writing,
        "waiting": latest_metric.frontend_waiting,
        "accepts": frontend_from_nginx.get("accepts", 0),
        "handled": frontend_from_nginx.get("handled", 0),
        "requests_total": frontend_from_nginx.get("requests_total", 0),
    }
    
    backend_instance = {
        "status": "ok",
        "cpu_percent": None,
        "memory_mb": latest_metric.backend_memory_mb,
        "connection_count": latest_metric.backend_active_requests,
        "errors": latest_metric.backend_errors_5xx,
        "discards": latest_metric.backend_discards_4xx,
        "details": {
            "requests_total": latest_metric.backend_requests_total,
            "avg_latency_ms": latest_metric.backend_avg_latency_ms,
            "db_pool_size": None,
            "db_pool_checked_out": None,
            "uptime_seconds": 0,
            "recent_errors": [],
        },
    }
    
    db_instance = {
        "status": "ok",
        "cpu_percent": None,
        "memory_mb": None,
        "connection_count": latest_metric.db_connections,
        "errors": latest_metric.db_errors,
        "discards": 0,
        "details": {
            "threads_running": latest_metric.db_threads_running,
            "questions_total": 0,
            "uptime_seconds": 0,
            "aborted_connects": 0,
            "connection_errors_max_connections": 0,
            "bytes_received": 0,
            "bytes_sent": 0,
            "slow_queries": 0,
            "select_full_join": 0,
            "tmp_disk_tables": 0,
        },
    }
    
    instance_logs = _derive_instance_logs(
        history=history,
        frontend_instance=frontend_instance,
        backend_instance=backend_instance,
        db_instance=db_instance,
        alerts=[],
        backend_recent_errors=[],
    )

    return {
        "generated_at": datetime.utcnow().isoformat(),
        "start_time": start.isoformat(),
        "end_time": end.isoformat(),
        "record_count": len(history),
        "instances": {
            "frontend": frontend_instance,
            "backend": backend_instance,
            "db": db_instance,
        },
        "alerts": [],
        "history": history,
        "instance_logs": instance_logs,
        "notes": [
            f"Data range: {len(history)} snapshots from historical database",
            "Instance metrics derived from latest snapshot in range",
        ],
    }


# ==================== Prometheus Metrics Endpoint ====================

def _prom_line(name: str, help_text: str, metric_type: str, value, labels: Optional[Dict[str, str]] = None) -> str:
    """Format a single Prometheus metric in text exposition format."""
    if value is None:
        return ""
    label_str = ""
    if labels:
        parts = ','.join(f'{k}="{v}"' for k, v in labels.items())
        label_str = f"{{{parts}}}"
    return f"# HELP {name} {help_text}\n# TYPE {name} {metric_type}\n{name}{label_str} {float(value)}\n"


@app.get("/metrics", include_in_schema=False)
def prometheus_metrics(db: Session = Depends(get_db)):
    """
    Prometheus-compatible metrics endpoint.
    Scrape this with Prometheus; visualise in Grafana.
    Note: restrict access to internal networks in production.
    """
    lines: List[str] = []

    # ---- Backend ----
    latency_samples = max(1, APP_INSIGHTS_STATE["backend_latency_samples"])
    avg_latency_ms = APP_INSIGHTS_STATE["backend_latency_ms_total"] / latency_samples
    uptime_seconds = int((datetime.utcnow() - APP_INSIGHTS_STATE["started_at"]).total_seconds())

    usage = resource.getrusage(resource.RUSAGE_SELF)
    backend_memory_mb = float(usage.ru_maxrss) / 1024.0
    if backend_memory_mb > 1024 * 1024:
        backend_memory_mb = float(usage.ru_maxrss) / (1024.0 * 1024.0)

    pool_size = None
    pool_checked_out = None
    try:
        if hasattr(engine.pool, "size"):
            pool_size = int(engine.pool.size())
        if hasattr(engine.pool, "checkedout"):
            pool_checked_out = int(engine.pool.checkedout())
    except Exception:
        pass

    lines.append(_prom_line("ledger_backend_active_requests", "Number of currently active HTTP requests", "gauge", APP_INSIGHTS_STATE["active_requests"]))
    lines.append(_prom_line("ledger_backend_requests_total", "Total HTTP requests handled since startup", "counter", APP_INSIGHTS_STATE["total_requests"]))
    lines.append(_prom_line("ledger_backend_errors_5xx_total", "Total 5xx server errors since startup", "counter", APP_INSIGHTS_STATE["errors_5xx"]))
    lines.append(_prom_line("ledger_backend_discards_4xx_total", "Total 4xx client errors since startup", "counter", APP_INSIGHTS_STATE["discards_4xx"]))
    lines.append(_prom_line("ledger_backend_latency_ms_avg", "Average backend request latency in milliseconds", "gauge", round(avg_latency_ms, 3)))
    lines.append(_prom_line("ledger_backend_memory_mb", "Backend process RSS memory usage in megabytes", "gauge", round(backend_memory_mb, 2)))
    lines.append(_prom_line("ledger_backend_uptime_seconds", "Backend process uptime in seconds", "counter", uptime_seconds))
    if pool_size is not None:
        lines.append(_prom_line("ledger_backend_db_pool_size", "SQLAlchemy connection pool size", "gauge", pool_size))
    if pool_checked_out is not None:
        lines.append(_prom_line("ledger_backend_db_pool_checked_out", "SQLAlchemy connections currently checked out", "gauge", pool_checked_out))

    # ---- Frontend (nginx stub_status) ----
    fe = _fetch_frontend_nginx_status()
    fe_up = 1 if fe.get("status") == "ok" else 0
    lines.append(_prom_line("ledger_frontend_up", "1 if nginx stub_status is reachable, 0 otherwise", "gauge", fe_up))
    if fe_up:
        lines.append(_prom_line("ledger_frontend_active_connections", "nginx active connections", "gauge", fe.get("connection_count")))
        lines.append(_prom_line("ledger_frontend_reading", "nginx connections in reading state", "gauge", fe.get("reading")))
        lines.append(_prom_line("ledger_frontend_writing", "nginx connections in writing state", "gauge", fe.get("writing")))
        lines.append(_prom_line("ledger_frontend_waiting", "nginx connections in waiting (keep-alive) state", "gauge", fe.get("waiting")))
        lines.append(_prom_line("ledger_frontend_accepts_total", "nginx total accepted connections", "counter", fe.get("accepts")))
        lines.append(_prom_line("ledger_frontend_handled_total", "nginx total handled connections", "counter", fe.get("handled")))
        lines.append(_prom_line("ledger_frontend_requests_total", "nginx total requests handled", "counter", fe.get("requests_total")))

    # ---- Database (MySQL GLOBAL STATUS) ----
    mysql_vars = [
        "Threads_connected", "Threads_running",
        "Aborted_clients", "Aborted_connects",
        "Connection_errors_max_connections",
        "Questions", "Uptime",
        "Bytes_received", "Bytes_sent",
        "Slow_queries", "Select_full_join", "Created_tmp_disk_tables",
    ]
    try:
        mysql_stats = _get_mysql_status_map(db, mysql_vars)
        db_errors = (
            mysql_stats.get("Aborted_clients", 0)
            + mysql_stats.get("Aborted_connects", 0)
            + mysql_stats.get("Connection_errors_max_connections", 0)
        )
        lines.append(_prom_line("ledger_db_connections", "MySQL Threads_connected (active client connections)", "gauge", mysql_stats.get("Threads_connected")))
        lines.append(_prom_line("ledger_db_threads_running", "MySQL Threads_running (threads actively executing queries)", "gauge", mysql_stats.get("Threads_running")))
        lines.append(_prom_line("ledger_db_errors_total", "MySQL cumulative connection/client errors", "counter", db_errors))
        lines.append(_prom_line("ledger_db_aborted_clients_total", "MySQL Aborted_clients count", "counter", mysql_stats.get("Aborted_clients")))
        lines.append(_prom_line("ledger_db_aborted_connects_total", "MySQL Aborted_connects count", "counter", mysql_stats.get("Aborted_connects")))
        lines.append(_prom_line("ledger_db_questions_total", "MySQL Questions (statements executed)", "counter", mysql_stats.get("Questions")))
        lines.append(_prom_line("ledger_db_uptime_seconds", "MySQL server uptime in seconds", "counter", mysql_stats.get("Uptime")))
        lines.append(_prom_line("ledger_db_bytes_received_total", "MySQL bytes received from clients", "counter", mysql_stats.get("Bytes_received")))
        lines.append(_prom_line("ledger_db_bytes_sent_total", "MySQL bytes sent to clients", "counter", mysql_stats.get("Bytes_sent")))
        lines.append(_prom_line("ledger_db_slow_queries_total", "MySQL slow queries count", "counter", mysql_stats.get("Slow_queries")))
        lines.append(_prom_line("ledger_db_select_full_join_total", "MySQL full-join selects (no index on join col)", "counter", mysql_stats.get("Select_full_join")))
        lines.append(_prom_line("ledger_db_tmp_disk_tables_total", "MySQL temporary tables created on disk", "counter", mysql_stats.get("Created_tmp_disk_tables")))
    except Exception:
        lines.append(_prom_line("ledger_db_up", "1 if DB metrics are available, 0 otherwise", "gauge", 0))

    # ---- Alerts summary ----
    # Re-derive alert counts from the current in-memory state (lightweight, no extra query)
    alert_counts: Dict[str, int] = {"high": 0, "medium": 0, "low": 0}
    if APP_INSIGHTS_STATE["errors_5xx"] > 0:
        alert_counts["high"] += 1
    if APP_INSIGHTS_STATE["discards_4xx"] > 20:
        alert_counts["medium"] += 1
    lines.append("# HELP ledger_alerts_active Number of active alerts by severity\n# TYPE ledger_alerts_active gauge\n")
    for severity, count in alert_counts.items():
        lines.append(f'ledger_alerts_active{{severity="{severity}"}} {count}\n')

    # ---- History snapshot count ----
    lines.append(_prom_line("ledger_insights_history_snapshots", "Number of in-memory App Insights snapshots retained", "gauge", len(APP_INSIGHTS_HISTORY)))

    body = "".join(line for line in lines if line)
    return PlainTextResponse(content=body, media_type="text/plain; version=0.0.4; charset=utf-8")


# ==================== Transaction Endpoints ====================

@app.get("/api/transactions", response_model=List[TransactionSchema])
def get_transactions(
    skip: int = 0, 
    limit: int = 5000,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all transactions for current user"""
    transactions = db.query(Transaction).filter(
        Transaction.user_id == current_user.id
    ).order_by(Transaction.date.desc()).offset(skip).limit(limit).all()
    return transactions


@app.post("/api/transactions/upload", response_model=List[TransactionSchema])
def upload_transactions_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Upload transactions from a CSV file.
    Required columns: type, description, amount, category, date
    Optional columns: notes, recurring (true/false), spread_over_year (true/false)
    """
    try:
        data = file.file.read().decode('utf-8-sig')
        csv_reader = csv.DictReader(io.StringIO(data))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read CSV file: {e}")

    if not csv_reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV file has no header")

    required_cols = {'type', 'description', 'amount', 'category', 'date'}
    missing = required_cols - set([c.strip() for c in csv_reader.fieldnames if c])
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(missing)}")

    created_transactions = []
    for idx, row in enumerate(csv_reader, start=1):
        if not row.get('type') or not row.get('description') or not row.get('amount') or not row.get('category') or not row.get('date'):
            continue

        tx_type = row.get('type', '').strip().lower()
        if tx_type not in ('income', 'expense'):
            raise HTTPException(status_code=400, detail=f"Invalid transaction type on row {idx}: {tx_type}")

        try:
            amount = float(row.get('amount'))
        except Exception:
            raise HTTPException(status_code=400, detail=f"Invalid amount on row {idx}: {row.get('amount')}")

        notes = row.get('notes', '').strip() if row.get('notes') else ''
        recurring_value = row.get('recurring', '').strip().lower()
        recurring = recurring_value in ('1', 'true', 'yes', 'y')
        spread_over_year_value = row.get('spread_over_year', '').strip().lower()
        spread_over_year = spread_over_year_value in ('1', 'true', 'yes', 'y')

        try:
            transaction_date = datetime.fromisoformat(row.get('date').strip()) if 'T' in row.get('date').strip() else datetime.strptime(row.get('date').strip(), '%Y-%m-%d')
        except Exception:
            raise HTTPException(status_code=400, detail=f"Invalid date format on row {idx}; expected YYYY-MM-DD or ISO format")

        db_transaction = Transaction(
            user_id=current_user.id,
            type=TransactionType(tx_type),
            description=row.get('description').strip(),
            amount=amount,
            category=row.get('category').strip(),
            date=transaction_date,
            notes=notes,
            recurring=recurring,
            spread_over_year=spread_over_year,
            synced=False,
            source='csv'
        )

        db.add(db_transaction)
        created_transactions.append(db_transaction)

    db.commit()

    for tx in created_transactions:
        db.refresh(tx)

    return created_transactions


@app.post("/api/transactions/upload-statement-file", response_model=StatementImportResponse)
@app.post("/api/transactions/upload-statement-pdf", response_model=StatementImportResponse)
def upload_statement_file(
    file: UploadFile = File(...),
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Upload a bank statement spreadsheet and import detected transactions."""
    extension = os.path.splitext(file.filename or "")[1].lower()
    if not file.filename or extension not in STATEMENT_FILE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Please upload a CSV, XLS, or XLSX bank statement")

    file_bytes = file.file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    stored_filename = f"{uuid4().hex}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, stored_filename)
    with open(file_path, "wb") as uploaded_file:
        uploaded_file.write(file_bytes)

    title = os.path.splitext(file.filename)[0] or "Statement Import"
    db_doc = Document(
        user_id=current_user.id,
        title=title,
        folder="Statements",
        subfolder="Imported Statements",
        file_name=file.filename,
        file_path=file_path,
        content_type=file.content_type,
        document_type="statement_file",
        frozen_import=False,
        imported_transaction_count=0,
    )
    db.add(db_doc)
    db.commit()
    db.refresh(db_doc)

    parse_message = "Statement file processed successfully"
    try:
        detected_transactions = _parse_statement_spreadsheet_transactions(file.filename, file_bytes)
    except HTTPException as exc:
        detected_transactions = []
        parse_message = f"Statement uploaded, but parsing did not complete: {exc.detail}"

    existing_transactions = db.query(Transaction).filter(Transaction.user_id == current_user.id).all()
    existing_fingerprints = set()
    for transaction in existing_transactions:
        normalized = {
            "date": transaction.date,
            "amount": transaction.amount,
            "type": transaction.type.value if hasattr(transaction.type, "value") else transaction.type,
            "category": transaction.category,
            "description": transaction.description,
        }
        existing_fingerprints.add(_statement_fingerprint(normalized))

    imported_transactions = []
    skipped_count = 0
    document_tag = _statement_document_note(db_doc.id)

    for entry in detected_transactions:
        fingerprint = _statement_fingerprint(entry)
        if fingerprint in existing_fingerprints:
            skipped_count += 1
            continue

        salary_cycle_tag = f" {STATEMENT_SALARY_NEXT_MONTH_TAG}" if _is_salary_credit_for_next_month(entry) else ""
        adjustment_tag = f" {STATEMENT_NON_INCOME_CREDIT_TAG}" if entry.get("is_credit_adjustment") else ""
        balance_reconciled_tag = f" {STATEMENT_BALANCE_RECONCILED_TAG}" if entry.get("balance_reconciled") else ""

        db_transaction = Transaction(
            user_id=current_user.id,
            type=TransactionType(str(entry["type"])),
            description=str(entry["description"]),
            amount=float(entry["amount"]),
            category=str(entry["category"]),
            date=entry["date"],
            notes=f"Imported from statement file ({file.filename}) {document_tag} [{fingerprint}]{salary_cycle_tag}{adjustment_tag}{balance_reconciled_tag}",
            recurring=False,
            spread_over_year=False,
            synced=False,
            source="statement_import"
        )
        db.add(db_transaction)
        imported_transactions.append(db_transaction)
        existing_fingerprints.add(fingerprint)

    liabilities_created = 0
    liabilities_updated = 0
    liability_candidates = [
        tx for tx in imported_transactions
        if (tx.type.value if hasattr(tx.type, "value") else str(tx.type)) == "expense"
        and (tx.category or "").strip() == "Loans (EMI)"
    ]
    if liability_candidates:
        existing_liabilities = db.query(Liability).filter(Liability.user_id == current_user.id).all()
        liability_map = {str(liability.lender or "").strip().lower(): liability for liability in existing_liabilities}

        for tx in liability_candidates:
            lender = _statement_lender_name(tx.description)
            lender_key = lender.lower()
            monthly_payment = float(tx.amount or 0)
            note_tag = f"Auto-created from statement file ({file.filename}) {document_tag}"

            existing_liability = liability_map.get(lender_key)
            if existing_liability:
                changed = False
                existing_monthly_payment = float(existing_liability.monthly_payment or 0)
                if monthly_payment > 0 and existing_monthly_payment <= 0:
                    existing_liability.monthly_payment = monthly_payment
                    changed = True
                if monthly_payment > 0 and float(existing_liability.outstanding or 0) <= 0:
                    existing_liability.outstanding = monthly_payment
                    changed = True
                if monthly_payment > 0 and float(existing_liability.amount or 0) <= 0:
                    existing_liability.amount = monthly_payment
                    changed = True
                if existing_liability.is_loan is not True:
                    existing_liability.is_loan = True
                    changed = True
                if not existing_liability.loan_start_date:
                    existing_liability.loan_start_date = tx.date
                    changed = True
                if changed:
                    liabilities_updated += 1
            else:
                created_liability = Liability(
                    user_id=current_user.id,
                    lender=lender,
                    amount=monthly_payment if monthly_payment > 0 else 0.0,
                    outstanding=monthly_payment if monthly_payment > 0 else 0.0,
                    is_loan=True,
                    loan_start_date=tx.date,
                    loan_tenure_months=None,
                    interest_rate=None,
                    opportunity_cost_rate=None,
                    monthly_payment=monthly_payment if monthly_payment > 0 else None,
                    linked_asset_id=None,
                    due_date=None,
                    notes=note_tag,
                )
                db.add(created_liability)
                liability_map[lender_key] = created_liability
                liabilities_created += 1

    db_doc.imported_transaction_count = len(imported_transactions)
    db.commit()

    response_transactions = []
    for transaction in imported_transactions:
        db.refresh(transaction)
        response_transactions.append(StatementImportTransaction(
            date=transaction.date.date().isoformat(),
            description=transaction.description,
            amount=transaction.amount,
            type=transaction.type.value if hasattr(transaction.type, "value") else str(transaction.type),
            category=transaction.category,
        ))

    return StatementImportResponse(
        message=parse_message if detected_transactions else f"{parse_message} No ledger rows were imported.",
        detected=len(detected_transactions),
        imported=len(imported_transactions),
        skipped=skipped_count,
        liabilities_created=liabilities_created,
        liabilities_updated=liabilities_updated,
        document=_serialize_document(db_doc),
        transactions=response_transactions,
    )


@app.post("/api/transactions", response_model=TransactionSchema)
def create_transaction(
    transaction: TransactionCreate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Create a new transaction"""
    db_transaction = Transaction(
        **transaction.dict(),
        user_id=current_user.id
    )
    db.add(db_transaction)
    db.commit()
    db.refresh(db_transaction)
    
    # Log audit
    log_audit(db, current_user.id, "create", "transaction", db_transaction.id,
              f"Created transaction: {transaction.description}",
              details=transaction.dict(),
              user_agent=request.headers.get("user-agent"))
    
    return db_transaction


@app.put("/api/transactions/{transaction_id}", response_model=TransactionSchema)
def update_transaction(
    transaction_id: int,
    transaction: TransactionUpdate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Update a transaction"""
    db_transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == current_user.id
    ).first()
    
    if not db_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if db_transaction.synced:
        raise HTTPException(status_code=400, detail="Cannot edit synced transactions")
    
    old_data = {
        "description": db_transaction.description,
        "amount": db_transaction.amount,
        "category": db_transaction.category
    }
    
    # Update fields
    try:
        update_data = transaction.dict(exclude_unset=True)

        # Replace string values and date field with parsed datetime in case incoming API sends date string
        if 'date' in update_data and isinstance(update_data['date'], str):
            from datetime import datetime
            update_data['date'] = datetime.fromisoformat(update_data['date'])

        for key, value in update_data.items():
            setattr(db_transaction, key, value)

        db.commit()
        db.refresh(db_transaction)

        # Log audit
        log_audit(db, current_user.id, "update", "transaction", transaction_id,
                  f"Updated transaction: {db_transaction.description}",
                  details={"old": old_data, "new": update_data},
                  user_agent=request.headers.get("user-agent"))

        return db_transaction
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Update transaction error: {type(exc).__name__}: {exc}")


@app.get("/api/assets", response_model=List[AssetSchema])
def get_assets(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    return db.query(Asset).filter(Asset.user_id == current_user.id).all()


@app.post("/api/assets", response_model=AssetSchema)
def create_asset(
    asset: AssetCreate,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    try:
        db_asset = Asset(**asset.dict(), user_id=current_user.id)
        db.add(db_asset)
        db.commit()
        db.refresh(db_asset)
        return db_asset
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Create asset error: {type(exc).__name__}: {exc}")


@app.put("/api/assets/{asset_id}", response_model=AssetSchema)
def update_asset(
    asset_id: int,
    asset: AssetUpdate,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    db_asset = db.query(Asset).filter(Asset.id == asset_id, Asset.user_id == current_user.id).first()
    if not db_asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    try:
        data = asset.dict(exclude_unset=True)
        for key, value in data.items():
            setattr(db_asset, key, value)

        db.commit()
        db.refresh(db_asset)
        return db_asset
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Update asset error: {type(exc).__name__}: {exc}")


@app.delete("/api/assets/{asset_id}")
def delete_asset(
    asset_id: int,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    db_asset = db.query(Asset).filter(Asset.id == asset_id, Asset.user_id == current_user.id).first()
    if not db_asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    db.delete(db_asset)
    db.commit()
    return {"message": "Asset deleted successfully"}


@app.post("/api/reports/expense-email", response_model=ExpenseReportEmailResponse)
def send_expense_report_email(
    payload: ExpenseReportEmailRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Generate a monthly expense report with charts and email it to the requested address."""
    month_start, month_end, month_label = _parse_report_month(payload.report_month)
    expenses = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.type == TransactionType.EXPENSE,
        Transaction.date >= month_start,
        Transaction.date <= month_end
    ).order_by(Transaction.date.asc()).all()

    report_html = _build_expense_report_html(current_user, month_label, expenses)
    gmail_integration = _get_gmail_integration(db, current_user.id)
    _send_gmail_message(
        db=db,
        integration=gmail_integration,
        recipient_email=payload.recipient_email,
        subject=f"Ledger Expense Report - {month_label}",
        html_body=report_html
    )

    return {
        "message": "Expense report sent successfully.",
        "recipient_email": payload.recipient_email,
        "report_month": month_start.strftime("%Y-%m"),
    }


@app.get("/api/data/export-csv")
def export_all_data_csv(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Export transactions, investments, liabilities, assets, budgets, goals, and summaries as CSV."""
    rows = _build_export_rows(current_user, db)
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=EXPORT_COLUMNS)
    writer.writeheader()
    for row in rows:
        normalized = {key: row.get(key, "") for key in EXPORT_COLUMNS}
        writer.writerow(normalized)

    filename = f"ledger-export-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.csv"
    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/api/data/import-csv", response_model=DataImportResponse)
def import_all_data_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Import a full Ledger export CSV and upsert supported entities."""
    try:
        raw_data = file.file.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(raw_data))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read CSV file: {exc}") from exc

    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV file has no header")

    created = 0
    updated = 0
    skipped = 0

    for row in reader:
        section = (row.get("section") or "").strip().lower()
        if not section or section == "summary":
            skipped += 1
            continue

        row_id = row.get("id")
        row_id = int(row_id) if row_id and row_id.strip().isdigit() else None

        if section == "transaction":
            raw_type = (row.get("type") or "expense").strip().lower()
            try:
                transaction_type = TransactionType(raw_type)
            except ValueError:
                skipped += 1
                continue

            payload = {
                "type": transaction_type,
                "description": row.get("description") or "",
                "amount": _csv_to_float(row.get("amount")),
                "category": row.get("category") or "Other",
                "date": _csv_to_datetime(row.get("date")),
                "notes": row.get("notes") or None,
                "recurring": _csv_to_bool(row.get("recurring")),
                "spread_over_year": _csv_to_bool(row.get("spread_over_year")),
                "source": row.get("source") or None,
            }
            if row_id:
                existing = db.query(Transaction).filter(Transaction.user_id == current_user.id, Transaction.id == row_id).first()
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    db.add(Transaction(id=row_id, user_id=current_user.id, synced=False, **payload))
                    created += 1
            else:
                db.add(Transaction(user_id=current_user.id, synced=False, **payload))
                created += 1

        elif section == "investment":
            payload = {
                "name": row.get("name") or "",
                "type": row.get("type") or "",
                "amount_invested": _csv_to_float(row.get("amount")),
                "current_value": _csv_to_float(row.get("current_value"), None) if row.get("current_value") not in ("", None) else None,
                "annual_growth_rate": _csv_to_float(row.get("annual_growth_rate"), None) if row.get("annual_growth_rate") not in ("", None) else None,
                "monthly_sip": _csv_to_bool(row.get("recurring")),
                "notes": row.get("notes") or None,
            }
            if row_id:
                existing = db.query(Investment).filter(Investment.user_id == current_user.id, Investment.id == row_id).first()
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    db.add(Investment(id=row_id, user_id=current_user.id, **payload))
                    created += 1
            else:
                db.add(Investment(user_id=current_user.id, **payload))
                created += 1

        elif section == "liability":
            payload = {
                "lender": row.get("lender") or "",
                "amount": _csv_to_float(row.get("amount")),
                "outstanding": _csv_to_float(row.get("outstanding")),
                "is_loan": _csv_to_bool(row.get("is_loan")),
                "loan_start_date": _csv_to_datetime(row.get("loan_start_date")) if row.get("loan_start_date") else None,
                "loan_tenure_months": int(row.get("loan_tenure_months")) if row.get("loan_tenure_months") not in ("", None) else None,
                "interest_rate": _csv_to_float(row.get("interest_rate"), None) if row.get("interest_rate") not in ("", None) else None,
                "opportunity_cost_rate": _csv_to_float(row.get("opportunity_cost_rate"), None) if row.get("opportunity_cost_rate") not in ("", None) else None,
                "monthly_payment": _csv_to_float(row.get("monthly_payment"), None) if row.get("monthly_payment") not in ("", None) else None,
                "linked_asset_id": int(row.get("linked_asset_id")) if row.get("linked_asset_id") not in ("", None) else None,
                "due_date": _csv_to_datetime(row.get("due_date")) if row.get("due_date") else None,
                "notes": row.get("notes") or None,
            }
            if row_id:
                existing = db.query(Liability).filter(Liability.user_id == current_user.id, Liability.id == row_id).first()
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    db.add(Liability(id=row_id, user_id=current_user.id, **payload))
                    created += 1
            else:
                db.add(Liability(user_id=current_user.id, **payload))
                created += 1

        elif section == "asset":
            payload = {
                "name": row.get("name") or "",
                "type": row.get("type") or None,
                "value": _csv_to_float(row.get("value")),
                "description": row.get("description") or None,
                "include_in_balance": _csv_to_bool(row.get("include_in_balance")),
                "include_in_income": _csv_to_bool(row.get("include_in_income")),
                "loan_emi_linked": _csv_to_bool(row.get("loan_emi_linked")),
            }
            if row_id:
                existing = db.query(Asset).filter(Asset.user_id == current_user.id, Asset.id == row_id).first()
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    db.add(Asset(id=row_id, user_id=current_user.id, **payload))
                    created += 1
            else:
                db.add(Asset(user_id=current_user.id, **payload))
                created += 1

        elif section == "budget":
            payload = {
                "category": row.get("category") or "Other",
                "limit": _csv_to_float(row.get("limit")),
                "period": row.get("period", "monthly").strip().lower() if row.get("period") else "monthly",
                "recurring": _csv_to_bool(row.get("recurring")),
                "start_month": row.get("start_month") or None,
            }
            # Validate period
            if payload["period"] not in ("monthly", "yearly"):
                payload["period"] = "monthly"
            
            if row_id:
                existing = db.query(Budget).filter(Budget.user_id == current_user.id, Budget.id == row_id).first()
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    db.add(Budget(id=row_id, user_id=current_user.id, **payload))
                    created += 1
            else:
                db.add(Budget(user_id=current_user.id, **payload))
                created += 1

        elif section == "goal":
            payload = {
                "name": row.get("name") or "",
                "target": _csv_to_float(row.get("target")),
                "current": _csv_to_float(row.get("current")),
                "target_date": _csv_to_datetime(row.get("target_date")) if row.get("target_date") else None,
            }
            if row_id:
                existing = db.query(Goal).filter(Goal.user_id == current_user.id, Goal.id == row_id).first()
                if existing:
                    for key, value in payload.items():
                        setattr(existing, key, value)
                    updated += 1
                else:
                    db.add(Goal(id=row_id, user_id=current_user.id, **payload))
                    created += 1
            else:
                db.add(Goal(user_id=current_user.id, **payload))
                created += 1
        else:
            skipped += 1

    db.commit()
    return {
        "message": "CSV import completed successfully.",
        "created": created,
        "updated": updated,
        "skipped": skipped,
    }


@app.get("/api/integrations/gmail/auth-url", response_model=IntegrationAuthUrlResponse)
def get_gmail_auth_url(
    request: Request,
    current_user: User = Depends(require_write_access)
):
    """Get a Google OAuth URL for connecting Gmail."""
    state = _create_google_oauth_state(current_user.id)
    params = {
        "client_id": _google_client_id(),
        "redirect_uri": _google_redirect_uri(request),
        "response_type": "code",
        "scope": " ".join(GOOGLE_GMAIL_SCOPES),
        "access_type": "offline",
        "prompt": "consent",
        "include_granted_scopes": "true",
        "state": state,
    }
    return {"auth_url": f"{GOOGLE_AUTH_BASE}?{urlencode(params)}"}


@app.get("/api/integrations/gmail/callback", response_class=HTMLResponse)
def gmail_oauth_callback(
    request: Request,
    code: Optional[str] = None,
    state: Optional[str] = None,
    error: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Handle Google OAuth callback for Gmail integration."""
    if error:
        return HTMLResponse(f"<html><body><script>window.opener&&window.opener.postMessage({{type:'gmail-connect-error',message:{json.dumps(error)}}}, '*');window.close();</script></body></html>")
    if not code or not state:
        raise HTTPException(status_code=400, detail="Missing OAuth callback parameters.")

    user_id = _decode_google_oauth_state(state)
    token_payload = _google_form_post(GOOGLE_TOKEN_URL, {
        "code": code,
        "client_id": _google_client_id(),
        "client_secret": _google_client_secret(),
        "redirect_uri": _google_redirect_uri(request),
        "grant_type": "authorization_code",
    })

    access_token = token_payload.get("access_token")
    if not access_token:
        raise HTTPException(status_code=500, detail="Google OAuth did not return an access token.")

    userinfo = _google_api_json(
        GOOGLE_USERINFO_URL,
        headers={"Authorization": f"Bearer {access_token}"},
    )
    token_payload["expires_at"] = (datetime.utcnow() + timedelta(seconds=int(token_payload.get("expires_in", 3600)) - 60)).isoformat()

    integration = _upsert_gmail_integration(
        db=db,
        user_id=user_id,
        token_data=token_payload,
        account_email=userinfo.get("email", ""),
    )

    html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; padding: 24px;">
        <p>Gmail connected successfully. You can close this window.</p>
        <script>
          if (window.opener) {{
            window.opener.postMessage({{
              type: 'gmail-connected',
              accountEmail: {json.dumps(integration.account_email or '')}
            }}, '*');
          }}
          window.close();
        </script>
      </body>
    </html>
    """
    return HTMLResponse(html)


# ---------------------------------------------------------------------------
# Gmail Bank Alert Sync
# ---------------------------------------------------------------------------

# Search query sent to Gmail to find bank transaction alert emails
_BANK_ALERT_GMAIL_QUERY = (
    "subject:(alert OR transaction OR debit OR credit OR \"bank statement\") "
    "from:(alerts hdfcbank icicibank sbi axisbank kotak paytm indusind yesbank "
    "rbl citi hsbc standard chartered idbi canara union punjab national "
    "amazon flipkart amex chase citibank notify noreply no-reply) "
    "newer_than:90d"
)


@app.post("/api/integrations/gmail/sync-bank-alerts", response_model=GmailBankAlertSyncResult)
def sync_gmail_bank_alerts(
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db),
):
    """
    Read bank alert emails from the connected Gmail account and import
    debits as expenses / credits as income into the transaction ledger.
    Duplicate detection: skips any email whose Gmail message-id was
    already imported (stored in transaction notes as gmail_msg_id:<id>).
    """
    integration = _get_gmail_integration(db, current_user.id)
    access_token = _get_valid_google_access_token(db, integration)

    # Build a set of already-imported Gmail message IDs for deduplication
    existing_notes = (
        db.query(Transaction.notes)
        .filter(
            Transaction.user_id == current_user.id,
            Transaction.source == "gmail",
            Transaction.notes.isnot(None),
        )
        .all()
    )
    imported_ids: set = set()
    for (note,) in existing_notes:
        if note and "gmail_msg_id:" in note:
            for part in note.split():
                if part.startswith("gmail_msg_id:"):
                    imported_ids.add(part[len("gmail_msg_id:"):])

    # Fetch message list from Gmail
    try:
        message_refs = _gmail_list_messages(access_token, _BANK_ALERT_GMAIL_QUERY, max_results=100)
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Failed to list Gmail messages: {exc}") from exc

    imported = 0
    skipped = 0
    errors = 0
    imported_txns = []

    for ref in message_refs:
        msg_id = ref.get("id")
        if not msg_id or msg_id in imported_ids:
            skipped += 1
            continue

        try:
            msg = _gmail_get_message(access_token, msg_id)
        except Exception:
            errors += 1
            continue

        payload = msg.get("payload", {})
        headers_list = payload.get("headers", [])
        header_map = {h["name"].lower(): h["value"] for h in headers_list}

        subject = header_map.get("subject", "")
        date_str = header_map.get("date", "")
        snippet = msg.get("snippet", "")

        # Decode body (text or html)
        body = _decode_gmail_body(payload) or snippet

        # Strip HTML tags if needed
        body_clean = re.sub(r"<[^>]+>", " ", body)
        body_clean = re.sub(r"&[a-zA-Z]+;", " ", body_clean)

        parsed = _parse_bank_alert_email(subject, body_clean)
        if not parsed:
            skipped += 1
            continue

        # Parse email date, fall back to now
        tx_date = datetime.utcnow()
        if date_str:
            for fmt in (
                "%a, %d %b %Y %H:%M:%S %z",
                "%d %b %Y %H:%M:%S %z",
                "%a, %d %b %Y %H:%M:%S %Z",
            ):
                try:
                    tx_date = datetime.strptime(date_str.strip(), fmt).replace(tzinfo=None)
                    break
                except ValueError:
                    continue

        tx_type = TransactionType.EXPENSE if parsed["type"] == "expense" else TransactionType.INCOME
        note_tag = f"gmail_msg_id:{msg_id}"

        tx = Transaction(
            user_id=current_user.id,
            type=tx_type,
            description=parsed["description"],
            amount=parsed["amount"],
            category=parsed["category"],
            date=tx_date,
            notes=note_tag,
            synced=True,
            source="gmail",
        )
        db.add(tx)
        imported_ids.add(msg_id)
        imported += 1
        imported_txns.append({
            "type": parsed["type"],
            "amount": parsed["amount"],
            "description": parsed["description"],
            "category": parsed["category"],
            "date": tx_date.isoformat(),
        })

    if imported > 0:
        integration.last_sync = datetime.utcnow()
        db.commit()

    log_audit(
        db=db,
        user_id=current_user.id,
        action="sync",
        entity_type="integration",
        entity_id="gmail",
        description=f"Gmail bank alert sync: {imported} imported, {skipped} skipped, {errors} errors",
        details={"imported": imported, "skipped": skipped, "errors": errors},
    )

    return GmailBankAlertSyncResult(
        message=f"Imported {imported} transaction(s) from bank alerts. {skipped} skipped, {errors} errors.",
        imported=imported,
        skipped=skipped,
        errors=errors,
        transactions=[
            {"type": t["type"], "amount": t["amount"], "description": t["description"],
             "category": t["category"], "date": t["date"]}
            for t in imported_txns
        ],
    )


@app.get("/api/documents", response_model=List[DocumentSchema])
def get_documents(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    docs = db.query(Document).filter(Document.user_id == current_user.id).all()
    return [_serialize_document(d) for d in docs]


@app.post("/api/documents", response_model=DocumentSchema)
def upload_document(
    title: str = Form(...),
    folder: str = Form("General"),
    subfolder: str = Form(""),
    file: UploadFile = File(...),
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    filename = f"{uuid4().hex}_{file.filename}"
    file_path = os.path.join(UPLOAD_DIR, filename)
    with open(file_path, "wb") as f:
        f.write(file.file.read())

    db_doc = Document(
        user_id=current_user.id,
        title=title,
        folder=folder.strip() or "General",
        subfolder=subfolder.strip() or None,
        file_name=file.filename,
        file_path=file_path,
        content_type=file.content_type,
        document_type="general",
        frozen_import=False,
        imported_transaction_count=0,
    )
    db.add(db_doc)
    db.commit()
    db.refresh(db_doc)

    return _serialize_document(db_doc)


@app.put("/api/documents/{document_id}", response_model=DocumentSchema)
def update_document(
    document_id: int,
    document_update: DocumentUpdate,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    db_doc = db.query(Document).filter(Document.id == document_id, Document.user_id == current_user.id).first()
    if not db_doc:
        raise HTTPException(status_code=404, detail="Document not found")

    update_data = document_update.model_dump(exclude_unset=True)

    if "title" in update_data:
        title = (update_data.get("title") or "").strip()
        if not title:
            raise HTTPException(status_code=400, detail="Document title is required")
        db_doc.title = title

    if "folder" in update_data:
        db_doc.folder = (update_data.get("folder") or "").strip() or "General"

    if "subfolder" in update_data:
        subfolder = update_data.get("subfolder")
        db_doc.subfolder = subfolder.strip() if isinstance(subfolder, str) and subfolder.strip() else None

    if "frozen_import" in update_data and _is_statement_document(db_doc):
        db_doc.frozen_import = bool(update_data.get("frozen_import"))

    db.commit()
    db.refresh(db_doc)

    return _serialize_document(db_doc)


@app.delete("/api/documents/{document_id}")
def delete_document(
    document_id: int,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    db_doc = db.query(Document).filter(Document.id == document_id, Document.user_id == current_user.id).first()
    if not db_doc:
        raise HTTPException(status_code=404, detail="Document not found")

    deleted_transactions = 0
    deleted_investments = 0
    deleted_liabilities = 0
    if _is_statement_document(db_doc) and not db_doc.frozen_import:
        statement_tag = _statement_document_note(db_doc.id)
        linked_transactions = db.query(Transaction).filter(
            Transaction.user_id == current_user.id,
            Transaction.notes.isnot(None),
            Transaction.notes.like(f"%{statement_tag}%")
        ).all()
        deleted_transactions = len(linked_transactions)
        for transaction in linked_transactions:
            db.delete(transaction)

        linked_investments = db.query(Investment).filter(
            Investment.user_id == current_user.id,
            Investment.notes.isnot(None),
            Investment.notes.like(f"%{statement_tag}%")
        ).all()
        deleted_investments = len(linked_investments)
        for investment in linked_investments:
            db.delete(investment)

        linked_liabilities = db.query(Liability).filter(
            Liability.user_id == current_user.id,
            Liability.notes.isnot(None),
            Liability.notes.like(f"%{statement_tag}%")
        ).all()
        deleted_liabilities = len(linked_liabilities)
        for liability in linked_liabilities:
            db.delete(liability)

    if os.path.exists(db_doc.file_path):
        os.remove(db_doc.file_path)

    db.delete(db_doc)
    db.commit()
    return {
        "message": "Document deleted successfully",
        "deleted_transactions": deleted_transactions,
        "deleted_investments": deleted_investments,
        "deleted_liabilities": deleted_liabilities,
        "frozen_import": bool(db_doc.frozen_import),
    }


@app.delete("/api/transactions/{transaction_id}")
def delete_transaction(
    transaction_id: int,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Delete a transaction"""
    db_transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == current_user.id
    ).first()
    
    if not db_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    if db_transaction.synced:
        raise HTTPException(status_code=400, detail="Cannot delete synced transactions")
    
    description = db_transaction.description
    db.delete(db_transaction)
    db.commit()
    
    # Log audit
    log_audit(db, current_user.id, "delete", "transaction", transaction_id,
              f"Deleted transaction: {description}",
              user_agent=request.headers.get("user-agent"))
    
    return {"message": "Transaction deleted successfully"}


# ==================== Budget Endpoints ====================

@app.get("/api/budgets", response_model=List[BudgetSchema])
def get_budgets(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all budgets for current user"""
    budgets = db.query(Budget).filter(Budget.user_id == current_user.id).all()
    return budgets


@app.post("/api/budgets", response_model=BudgetSchema)
def create_budget(
    budget: BudgetCreate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Create a new budget"""
    payload = budget.dict()
    if not payload.get("start_month"):
        payload["start_month"] = datetime.utcnow().strftime("%Y-%m")
    db_budget = Budget(**payload, user_id=current_user.id)
    db.add(db_budget)
    db.commit()
    db.refresh(db_budget)
    
    log_audit(db, current_user.id, "create", "budget", db_budget.id,
              f"Created budget for {budget.category}",
              details=budget.dict(),
              user_agent=request.headers.get("user-agent"))
    
    return db_budget


@app.put("/api/budgets/{budget_id}", response_model=BudgetSchema)
def update_budget(
    budget_id: int,
    budget_update: BudgetUpdate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Update an existing budget"""
    db_budget = db.query(Budget).filter(
        Budget.user_id == current_user.id,
        Budget.id == budget_id
    ).first()

    if not db_budget:
        raise HTTPException(status_code=404, detail="Budget not found")

    for field, value in budget_update.dict(exclude_unset=True).items():
        setattr(db_budget, field, value)

    db.commit()
    db.refresh(db_budget)

    log_audit(db, current_user.id, "update", "budget", db_budget.id,
              f"Updated budget for {db_budget.category}",
              details=budget_update.dict(exclude_unset=True),
              user_agent=request.headers.get("user-agent"))

    return db_budget


@app.delete("/api/budgets/{budget_id}", response_model=BudgetSchema)
def delete_budget(
    budget_id: int,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Delete a budget"""
    db_budget = db.query(Budget).filter(
        Budget.user_id == current_user.id,
        Budget.id == budget_id
    ).first()

    if not db_budget:
        raise HTTPException(status_code=404, detail="Budget not found")

    db.delete(db_budget)
    db.commit()

    log_audit(db, current_user.id, "delete", "budget", budget_id,
              f"Deleted budget for {db_budget.category}",
              user_agent=request.headers.get("user-agent"))

    return db_budget


@app.get("/api/budgets/spending/monthly", response_model=List[BudgetWithSpending])
def get_budgets_with_spending(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all budgets with current month/year spending details based on budget period"""
    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    budgets = db.query(Budget).filter(Budget.user_id == current_user.id).all()
    
    budgets_with_spending = []
    for budget in budgets:
        # Determine period for spending calculation
        period = getattr(budget, 'period', 'monthly') or 'monthly'
        
        if period == 'yearly':
            # Calculate yearly spending (from Jan 1 to now)
            start_date = start_of_year
        else:
            # Calculate monthly spending (from 1st of month to now)
            start_date = start_of_month
        
        # Calculate spending for this category in the specified period
        category_spending = db.query(Transaction).filter(
            Transaction.user_id == current_user.id,
            Transaction.type == "expense",
            Transaction.category == budget.category,
            Transaction.date >= start_date
        ).all()
        
        spent = sum(t.amount for t in category_spending)
        remaining = budget.limit - spent
        percentage_used = (spent / budget.limit * 100) if budget.limit > 0 else 0
        is_over_budget = spent > budget.limit
        
        budget_with_spending = BudgetWithSpending(
            id=budget.id,
            user_id=budget.user_id,
            category=budget.category,
            limit=budget.limit,
            period=period,
            spent=round(spent, 2),
            remaining=round(remaining, 2),
            percentage_used=round(percentage_used, 1),
            is_over_budget=is_over_budget,
            created_at=budget.created_at,
            updated_at=budget.updated_at
        )
        budgets_with_spending.append(budget_with_spending)
    
    return budgets_with_spending


# ==================== Goal Endpoints ====================

@app.get("/api/goals", response_model=List[GoalSchema])
def get_goals(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all goals for current user"""
    goals = db.query(Goal).filter(Goal.user_id == current_user.id).all()
    return goals


@app.post("/api/goals", response_model=GoalSchema)
def create_goal(
    goal: GoalCreate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Create a new goal"""
    db_goal = Goal(**goal.dict(), user_id=current_user.id)
    db.add(db_goal)
    db.commit()
    db.refresh(db_goal)
    
    log_audit(db, current_user.id, "create", "goal", db_goal.id,
              f"Created goal: {goal.name}",
              details=goal.dict(),
              user_agent=request.headers.get("user-agent"))
    
    return db_goal


@app.put("/api/goals/{goal_id}", response_model=GoalSchema)
def update_goal(
    goal_id: int,
    goal_update: GoalUpdate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Update an existing goal"""
    db_goal = db.query(Goal).filter(
        Goal.user_id == current_user.id,
        Goal.id == goal_id
    ).first()

    if not db_goal:
        raise HTTPException(status_code=404, detail="Goal not found")

    for field, value in goal_update.dict(exclude_unset=True).items():
        setattr(db_goal, field, value)

    db.commit()
    db.refresh(db_goal)

    log_audit(db, current_user.id, "update", "goal", db_goal.id,
              f"Updated goal: {db_goal.name}",
              details=goal_update.dict(exclude_unset=True),
              user_agent=request.headers.get("user-agent"))

    return db_goal


@app.delete("/api/goals/{goal_id}", response_model=GoalSchema)
def delete_goal(
    goal_id: int,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Delete a goal"""
    db_goal = db.query(Goal).filter(
        Goal.user_id == current_user.id,
        Goal.id == goal_id
    ).first()

    if not db_goal:
        raise HTTPException(status_code=404, detail="Goal not found")

    db.delete(db_goal)
    db.commit()

    log_audit(db, current_user.id, "delete", "goal", goal_id,
              f"Deleted goal: {db_goal.name}",
              user_agent=request.headers.get("user-agent"))

    return db_goal


# ==================== Investment Endpoints ====================

def _ensure_goal_belongs_to_user(db: Session, user_id: int, goal_id: int) -> None:
    """Validate that a referenced goal exists and belongs to the current user."""
    goal = db.query(Goal).filter(Goal.user_id == user_id, Goal.id == goal_id).first()
    if not goal:
        raise HTTPException(status_code=404, detail="Linked goal not found")


def _normalize_investment_payload(
    payload: dict,
    db: Session,
    user_id: int,
    existing: Optional[Investment] = None,
) -> dict:
    """Normalize and validate investment payload consistently for create/update paths."""
    normalized = dict(payload)

    investment_type = (normalized.get("type") or (existing.type if existing else "") or "").strip().lower()
    if "mutual fund" in investment_type and normalized.get("current_value") is None:
        fallback_amount = normalized.get("amount_invested")
        if fallback_amount is None and existing is not None:
            fallback_amount = existing.amount_invested
        normalized["current_value"] = fallback_amount

    effective_goal_id = normalized.get("goal_id")
    if effective_goal_id is None and existing is not None:
        effective_goal_id = existing.goal_id

    if effective_goal_id is None:
        raise HTTPException(status_code=400, detail="goal_id is required for investments")

    try:
        effective_goal_id = int(effective_goal_id)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="goal_id must be a valid integer")

    _ensure_goal_belongs_to_user(db, user_id, effective_goal_id)
    normalized["goal_id"] = effective_goal_id

    return normalized

@app.get("/api/investments", response_model=List[InvestmentSchema])
def get_investments(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all investments for current user"""
    return db.query(Investment).filter(Investment.user_id == current_user.id).all()


@app.post("/api/investments", response_model=InvestmentSchema)
def create_investment(
    investment: InvestmentCreate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Create a new investment record"""
    payload = _normalize_investment_payload(
        investment.dict(),
        db=db,
        user_id=current_user.id,
    )

    db_investment = Investment(**payload, user_id=current_user.id)
    db.add(db_investment)
    db.commit()
    db.refresh(db_investment)
    
    log_audit(db, current_user.id, "create", "investment", db_investment.id,
              f"Added investment: {investment.name}",
              details=investment.dict(),
              user_agent=request.headers.get("user-agent"))
    
    return db_investment


@app.put("/api/investments/{investment_id}", response_model=InvestmentSchema)
def update_investment(
    investment_id: int,
    investment_update: InvestmentUpdate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Update an existing investment"""
    db_investment = db.query(Investment).filter(
        Investment.user_id == current_user.id,
        Investment.id == investment_id
    ).first()

    if not db_investment:
        raise HTTPException(status_code=404, detail="Investment not found")

    update_data = _normalize_investment_payload(
        investment_update.dict(exclude_unset=True),
        db=db,
        user_id=current_user.id,
        existing=db_investment,
    )

    for field, value in update_data.items():
        setattr(db_investment, field, value)

    db.commit()
    db.refresh(db_investment)

    log_audit(db, current_user.id, "update", "investment", db_investment.id,
              f"Updated investment: {db_investment.name}",
              details=investment_update.dict(exclude_unset=True),
              user_agent=request.headers.get("user-agent"))

    return db_investment


@app.delete("/api/investments/{investment_id}", response_model=InvestmentSchema)
def delete_investment(
    investment_id: int,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Delete an investment record"""
    db_investment = db.query(Investment).filter(
        Investment.user_id == current_user.id,
        Investment.id == investment_id
    ).first()

    if not db_investment:
        raise HTTPException(status_code=404, detail="Investment not found")

    db.delete(db_investment)
    db.commit()

    log_audit(db, current_user.id, "delete", "investment", investment_id,
              f"Deleted investment: {db_investment.name}",
              user_agent=request.headers.get("user-agent"))

    return db_investment


# ==================== Liability Endpoints ====================

@app.get("/api/liabilities", response_model=List[LiabilitySchema])
def get_liabilities(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all liabilities for current user"""
    return db.query(Liability).filter(Liability.user_id == current_user.id).all()


@app.post("/api/liabilities", response_model=LiabilitySchema)
def create_liability(
    liability: LiabilityCreate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Create a new liability record"""
    db_liability = Liability(**liability.dict(), user_id=current_user.id)
    db.add(db_liability)
    db.commit()
    db.refresh(db_liability)

    log_audit(db, current_user.id, "create", "liability", db_liability.id,
              f"Added liability: {liability.lender}",
              details=liability.dict(),
              user_agent=request.headers.get("user-agent"))

    return db_liability


@app.put("/api/liabilities/{liability_id}", response_model=LiabilitySchema)
def update_liability(
    liability_id: int,
    liability_update: LiabilityUpdate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Update an existing liability"""
    db_liability = db.query(Liability).filter(
        Liability.user_id == current_user.id,
        Liability.id == liability_id
    ).first()

    if not db_liability:
        raise HTTPException(status_code=404, detail="Liability not found")

    for field, value in liability_update.dict(exclude_unset=True).items():
        setattr(db_liability, field, value)

    db.commit()
    db.refresh(db_liability)

    log_audit(db, current_user.id, "update", "liability", db_liability.id,
              f"Updated liability: {db_liability.lender}",
              details=liability_update.dict(exclude_unset=True),
              user_agent=request.headers.get("user-agent"))

    return db_liability


@app.delete("/api/liabilities/{liability_id}", response_model=LiabilitySchema)
def delete_liability(
    liability_id: int,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Delete a liability record"""
    db_liability = db.query(Liability).filter(
        Liability.user_id == current_user.id,
        Liability.id == liability_id
    ).first()

    if not db_liability:
        raise HTTPException(status_code=404, detail="Liability not found")

    db.delete(db_liability)
    db.commit()

    log_audit(db, current_user.id, "delete", "liability", liability_id,
              f"Deleted liability: {db_liability.lender}",
              user_agent=request.headers.get("user-agent"))

    return db_liability


# ==================== Integration Endpoints ====================

@app.get("/api/integrations", response_model=List[IntegrationSchema])
def get_integrations(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all integrations for current user"""
    integrations = db.query(Integration).filter(
        Integration.user_id == current_user.id
    ).all()
    return integrations


@app.post("/api/integrations", response_model=IntegrationSchema)
def create_integration(
    integration: IntegrationCreate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Create or update an integration"""
    # Check if integration exists
    db_integration = db.query(Integration).filter(
        Integration.user_id == current_user.id,
        Integration.app_name == integration.app_name
    ).first()
    
    if db_integration:
        # Update existing
        db_integration.connected = True
        db_integration.api_key = integration.api_key
        db_integration.sync_frequency = integration.sync_frequency
        db_integration.last_sync = datetime.utcnow()
        action = "update"
    else:
        # Create new
        db_integration = Integration(
            **integration.dict(),
            user_id=current_user.id,
            connected=True,
            last_sync=datetime.utcnow()
        )
        db.add(db_integration)
        action = "create"
    
    db.commit()
    db.refresh(db_integration)
    
    log_audit(db, current_user.id, action, "integration", db_integration.id,
              f"Connected {integration.app_name}",
              user_agent=request.headers.get("user-agent"))
    
    return db_integration


@app.put("/api/integrations/{app_name}", response_model=IntegrationSchema)
def update_integration(
    app_name: str,
    integration_update: IntegrationUpdate,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Update an existing integration (e.g., disconnect or update credentials)."""
    db_integration = db.query(Integration).filter(
        Integration.user_id == current_user.id,
        Integration.app_name == app_name
    ).first()

    if not db_integration:
        raise HTTPException(status_code=404, detail="Integration not found")

    if integration_update.api_key is not None:
        db_integration.api_key = integration_update.api_key
    if integration_update.account_email is not None:
        db_integration.account_email = integration_update.account_email
    if integration_update.sync_frequency is not None:
        db_integration.sync_frequency = integration_update.sync_frequency
    if integration_update.connected is not None:
        db_integration.connected = integration_update.connected

    if app_name == "gmail" and integration_update.connected is False:
        db_integration.oauth_token = None
        db_integration.account_email = None
        db_integration.api_key = None

    if db_integration.connected:
        db_integration.last_sync = datetime.utcnow()

    db.commit()
    db.refresh(db_integration)

    log_audit(db, current_user.id, "update", "integration", db_integration.id,
              f"Updated integration {app_name}",
              user_agent=request.headers.get("user-agent"))

    return db_integration


@app.delete("/api/integrations/{app_name}", response_model=IntegrationSchema)
def delete_integration(
    app_name: str,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Delete an integration configuration."""
    integration = db.query(Integration).filter(
        Integration.user_id == current_user.id,
        Integration.app_name == app_name
    ).first()

    if not integration:
        raise HTTPException(status_code=404, detail="Integration not found")

    db.delete(integration)
    db.commit()

    log_audit(db, current_user.id, "delete", "integration", app_name,
              f"Deleted integration {app_name}",
              user_agent=request.headers.get("user-agent"))

    return integration


@app.post("/api/integrations/{app_name}/sync")
def sync_integration(
    app_name: str,
    request: Request,
    current_user: User = Depends(require_write_access),
    db: Session = Depends(get_db)
):
    """Sync transactions from an integration"""
    integration = db.query(Integration).filter(
        Integration.user_id == current_user.id,
        Integration.app_name == app_name
    ).first()

    if not integration or not integration.connected:
        raise HTTPException(status_code=404, detail="Integration not found or not connected")

    # Use a provider-based integration system.
    provider = get_provider(integration, current_user)

    try:
        fetched_transactions = provider.fetch_transactions()
    except Exception as e:
        # Mark integration as disconnected on failure so UI reflects connection issues
        integration.connected = False
        db.commit()
        raise HTTPException(status_code=502, detail=str(e))

    synced_transactions = []
    for trans_data in fetched_transactions:
        # Normalize incoming transaction fields
        tx_type = trans_data.get("type") or trans_data.get("transaction_type") or "expense"
        tx_date = trans_data.get("date") or trans_data.get("timestamp")

        # Ensure date is parsable (fallback to now)
        try:
            parsed_date = datetime.fromisoformat(tx_date) if tx_date else datetime.utcnow()
        except Exception:
            parsed_date = datetime.utcnow()

        db_transaction = Transaction(
            user_id=current_user.id,
            synced=True,
            source=app_name,
            date=parsed_date,
            type=tx_type,
            description=trans_data.get("description") or "Imported transaction",
            amount=float(trans_data.get("amount") or 0),
            category=trans_data.get("category") or "Imported",
            notes=trans_data.get("notes") or None
        )
        db.add(db_transaction)
        synced_transactions.append(db_transaction)

    integration.last_sync = datetime.utcnow()
    db.commit()

    log_audit(db, current_user.id, "sync", "transaction", app_name,
              f"Synced {len(synced_transactions)} transactions from {app_name}",
              details={"count": len(synced_transactions)},
              user_agent=request.headers.get("user-agent"))

    return {"message": f"Synced {len(synced_transactions)} transactions", "count": len(synced_transactions)}


# ==================== Audit Log Endpoints ====================

@app.get("/api/audit-logs", response_model=List[AuditLogSchema])
def get_audit_logs(
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get audit logs for current user"""
    logs = db.query(AuditLog).filter(
        AuditLog.user_id == current_user.id
    ).order_by(AuditLog.timestamp.desc()).offset(skip).limit(limit).all()
    return logs


# ==================== Dashboard Endpoints ====================

@app.get("/api/dashboard/stats", response_model=DashboardStats)
def get_dashboard_stats(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get dashboard statistics"""
    # Get current month transactions
    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    transactions = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.date >= start_of_month
    ).all()
    
    income = sum(t.amount for t in transactions if t.type == "income")
    expenses = sum(t.amount for t in transactions if t.type == "expense")
    balance = income - expenses
    savings_rate = (balance / income * 100) if income > 0 else 0
    
    budget_count = db.query(Budget).filter(Budget.user_id == current_user.id).count()
    goal_count = db.query(Goal).filter(Goal.user_id == current_user.id).count()
    
    # Calculate budget spending
    budgets = db.query(Budget).filter(Budget.user_id == current_user.id).all()
    
    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    
    budgets_with_spending_list = []
    total_budget_limit = 0.0
    total_budget_spent = 0.0
    budgets_over_limit = 0
    
    for budget in budgets:
        # Determine period for spending calculation
        period = getattr(budget, 'period', 'monthly') or 'monthly'
        
        if period == 'yearly':
            start_date = start_of_year
        else:
            start_date = start_of_month
        
        # Calculate spending for this category in the specified period
        category_spending = db.query(Transaction).filter(
            Transaction.user_id == current_user.id,
            Transaction.type == "expense",
            Transaction.category == budget.category,
            Transaction.date >= start_date
        ).all()
        
        spent = sum(t.amount for t in category_spending)
        remaining = budget.limit - spent
        percentage_used = (spent / budget.limit * 100) if budget.limit > 0 else 0
        is_over_budget = spent > budget.limit
        
        budget_with_spending = BudgetWithSpending(
            id=budget.id,
            user_id=budget.user_id,
            category=budget.category,
            limit=budget.limit,
            period=period,
            spent=round(spent, 2),
            remaining=round(remaining, 2),
            percentage_used=round(percentage_used, 1),
            is_over_budget=is_over_budget,
            created_at=budget.created_at,
            updated_at=budget.updated_at
        )
        budgets_with_spending_list.append(budget_with_spending)
        
        total_budget_limit += budget.limit
        total_budget_spent += spent
        if is_over_budget:
            budgets_over_limit += 1
    
    budget_remaining = total_budget_limit - total_budget_spent
    
    return DashboardStats(
        total_balance=balance,
        total_income=income,
        total_expenses=expenses,
        savings_rate=round(savings_rate, 1),
        transaction_count=len(transactions),
        budget_count=budget_count,
        goal_count=goal_count,
        budgets_with_spending=budgets_with_spending_list,
        total_budget_limit=round(total_budget_limit, 2),
        total_budget_spent=round(total_budget_spent, 2),
        budget_remaining=round(budget_remaining, 2),
        budgets_over_limit=budgets_over_limit
    )


@app.get("/api/financial-insights")
def get_financial_insights(
    months: Optional[int] = 6,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Financial metrics and insights for the authenticated user.

    Returns income/expense trend, savings rate, net worth breakdown,
    budget utilisation, top spending categories, investment summary,
    and goal progress for the last `months` calendar months (default 6).
    """
    now = datetime.utcnow()
    start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # ── Current month ─────────────────────────────────────────────────────────
    cur_tx = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.date >= start_of_month,
    ).all()
    cur_income = sum(t.amount for t in cur_tx if t.type == "income")
    cur_expenses = sum(t.amount for t in cur_tx if t.type == "expense")
    cur_net = cur_income - cur_expenses
    savings_rate = round(cur_net / cur_income * 100, 2) if cur_income > 0 else 0.0

    # ── Previous month ────────────────────────────────────────────────────────
    prev_start = (start_of_month - timedelta(days=1)).replace(
        day=1, hour=0, minute=0, second=0, microsecond=0
    )
    prev_tx = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.date >= prev_start,
        Transaction.date < start_of_month,
    ).all()
    prev_income = sum(t.amount for t in prev_tx if t.type == "income")
    prev_expenses = sum(t.amount for t in prev_tx if t.type == "expense")

    # ── Monthly trend (last N months) ─────────────────────────────────────────
    trend = []
    for i in range(max(1, months) - 1, -1, -1):
        # Approximate month start by stepping back 28-day chunks and normalising
        m_start = (now.replace(day=1) - timedelta(days=i * 28)).replace(
            day=1, hour=0, minute=0, second=0, microsecond=0
        )
        # Month end = first day of the following month
        next_month_day = m_start.replace(day=28) + timedelta(days=4)
        m_end = next_month_day.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        m_tx = db.query(Transaction).filter(
            Transaction.user_id == current_user.id,
            Transaction.date >= m_start,
            Transaction.date < m_end,
        ).all()
        m_income = sum(t.amount for t in m_tx if t.type == "income")
        m_expenses = sum(t.amount for t in m_tx if t.type == "expense")
        m_net = m_income - m_expenses
        trend.append({
            "month": m_start.strftime("%b %Y"),
            "income": round(m_income, 2),
            "expenses": round(m_expenses, 2),
            "net": round(m_net, 2),
            "savings_rate_pct": round(m_net / m_income * 100, 2) if m_income > 0 else 0.0,
        })

    # ── Top spending categories (current month) ───────────────────────────────
    category_spend: Dict[str, float] = {}
    for t in cur_tx:
        if t.type == "expense":
            key = t.category or "Uncategorized"
            category_spend[key] = category_spend.get(key, 0.0) + t.amount
    top_categories = sorted(
        [{"category": k, "amount": round(v, 2)} for k, v in category_spend.items()],
        key=lambda x: x["amount"],
        reverse=True,
    )[:10]

    # ── Budget utilisation ────────────────────────────────────────────────────
    budgets = db.query(Budget).filter(Budget.user_id == current_user.id).all()
    budget_items = []
    total_budget_limit = 0.0
    total_budget_spent = 0.0
    over_budget_count = 0
    start_of_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    for b in budgets:
        period = (getattr(b, "period", "monthly") or "monthly").lower()
        since = start_of_year if period == "yearly" else start_of_month
        spent = sum(
            t.amount
            for t in db.query(Transaction).filter(
                Transaction.user_id == current_user.id,
                Transaction.type == "expense",
                Transaction.category == b.category,
                Transaction.date >= since,
            ).all()
        )
        pct = round(spent / b.limit * 100, 1) if b.limit > 0 else 0.0
        over = spent > b.limit
        budget_items.append({
            "category": b.category,
            "limit": round(b.limit, 2),
            "spent": round(spent, 2),
            "remaining": round(b.limit - spent, 2),
            "utilization_pct": pct,
            "over_budget": over,
        })
        total_budget_limit += b.limit
        total_budget_spent += spent
        if over:
            over_budget_count += 1

    # ── Investments ───────────────────────────────────────────────────────────
    investments = db.query(Investment).filter(Investment.user_id == current_user.id).all()
    total_invested = sum(float(i.amount_invested or 0) for i in investments)
    total_current_value = sum(float(i.current_value or i.amount_invested or 0) for i in investments)
    inv_by_type: Dict[str, float] = {}
    for i in investments:
        itype = i.type or "Other"
        inv_by_type[itype] = inv_by_type.get(itype, 0.0) + float(i.amount_invested or 0)

    # ── Assets & Liabilities ──────────────────────────────────────────────────
    assets = db.query(Asset).filter(Asset.user_id == current_user.id).all()
    total_assets = sum(float(a.value or 0) for a in assets)
    liabilities = db.query(Liability).filter(Liability.user_id == current_user.id).all()
    total_liabilities = sum(float(l.outstanding or 0) for l in liabilities)
    net_worth = total_assets + total_current_value - total_liabilities

    # ── Goals ─────────────────────────────────────────────────────────────────
    goals = db.query(Goal).filter(Goal.user_id == current_user.id).all()
    goal_items = []
    for g in goals:
        target = float(g.target or 0)
        current = float(g.current or 0)
        goal_items.append({
            "name": g.name,
            "target": round(target, 2),
            "current": round(current, 2),
            "progress_pct": round(current / target * 100, 1) if target > 0 else 0.0,
            "deadline": g.target_date.isoformat() if g.target_date else None,
        })

    # ── Recommendations ───────────────────────────────────────────────────────
    recommendations = []
    if savings_rate < 20:
        recommendations.append({
            "type": "warning",
            "message": f"Savings rate is {savings_rate:.1f}%. Target at least 20% of income.",
        })
    if over_budget_count > 0:
        recommendations.append({
            "type": "alert",
            "message": f"{over_budget_count} budget(s) exceeded this month.",
        })
    if cur_expenses > cur_income and cur_income > 0:
        recommendations.append({
            "type": "alert",
            "message": "Expenses exceed income this month — you are running a deficit.",
        })
    if net_worth < 0:
        recommendations.append({
            "type": "alert",
            "message": "Net worth is negative. Focus on reducing outstanding liabilities.",
        })
    if total_invested == 0:
        recommendations.append({
            "type": "info",
            "message": "No investments tracked yet. Consider starting an SIP or building a portfolio.",
        })
    if not recommendations:
        recommendations.append({
            "type": "ok",
            "message": "Finances look healthy this month. Keep it up!",
        })

    return {
        "generated_at": now.isoformat(),
        "period": {
            "month": now.strftime("%B %Y"),
            "months_included": months,
            "start": start_of_month.isoformat(),
        },
        "summary": {
            "current_month_income": round(cur_income, 2),
            "current_month_expenses": round(cur_expenses, 2),
            "current_month_net": round(cur_net, 2),
            "savings_rate_pct": savings_rate,
            "prev_month_income": round(prev_income, 2),
            "prev_month_expenses": round(prev_expenses, 2),
            "income_change_pct": round((cur_income - prev_income) / prev_income * 100, 2) if prev_income > 0 else None,
            "expenses_change_pct": round((cur_expenses - prev_expenses) / prev_expenses * 100, 2) if prev_expenses > 0 else None,
        },
        "net_worth": {
            "total_assets": round(total_assets, 2),
            "total_liabilities": round(total_liabilities, 2),
            "net_worth": round(net_worth, 2),
            "total_invested": round(total_invested, 2),
            "total_current_value": round(total_current_value, 2),
            "unrealised_gain": round(total_current_value - total_invested, 2),
        },
        "top_categories": top_categories,
        "monthly_trend": trend,
        "budget_utilization": {
            "total_limit": round(total_budget_limit, 2),
            "total_spent": round(total_budget_spent, 2),
            "total_remaining": round(total_budget_limit - total_budget_spent, 2),
            "over_budget_count": over_budget_count,
            "items": budget_items,
        },
        "investments": {
            "total_invested": round(total_invested, 2),
            "total_current_value": round(total_current_value, 2),
            "unrealised_gain": round(total_current_value - total_invested, 2),
            "by_type": {k: round(v, 2) for k, v in inv_by_type.items()},
            "count": len(investments),
            "sip_count": sum(1 for i in investments if i.monthly_sip),
        },
        "goals": goal_items,
        "recommendations": recommendations,
    }


# ==================== App Insights Historical Recording ====================

insights_recording_thread = None
insights_recording_active = False


def record_insights_snapshot():
    """Record a snapshot of current app insights metrics to the database."""
    try:
        db = next(get_db())
        
        # Get current metrics
        latency_samples = max(1, APP_INSIGHTS_STATE["backend_latency_samples"])
        avg_latency_ms = APP_INSIGHTS_STATE["backend_latency_ms_total"] / latency_samples
        
        # Get frontend nginx status
        frontend_status = _fetch_frontend_nginx_status()
        
        # Get DB metrics
        mysql_vars = [
            "Threads_connected", "Threads_running", "Aborted_clients",
            "Aborted_connects", "Connection_errors_max_connections", "Questions"
        ]
        mysql_stats = _get_mysql_status_map(db, mysql_vars)
        db_errors = (
            mysql_stats.get("Aborted_clients", 0) +
            mysql_stats.get("Aborted_connects", 0) +
            mysql_stats.get("Connection_errors_max_connections", 0)
        )
        
        # Create and save the metric record
        metric = AppInsightsMetric(
            recorded_at=datetime.utcnow(),
            # Backend metrics
            backend_active_requests=APP_INSIGHTS_STATE["active_requests"],
            backend_errors_5xx=APP_INSIGHTS_STATE["errors_5xx"],
            backend_discards_4xx=APP_INSIGHTS_STATE["discards_4xx"],
            backend_avg_latency_ms=round(avg_latency_ms, 2),
            backend_requests_total=APP_INSIGHTS_STATE["total_requests"],
            backend_memory_mb=0.0,  # Will update below
            # Frontend metrics
            frontend_active_connections=frontend_status.get("connection_count", 0),
            frontend_reading=frontend_status.get("reading", 0),
            frontend_writing=frontend_status.get("writing", 0),
            frontend_waiting=frontend_status.get("waiting", 0),
            frontend_accepts=frontend_status.get("accepts", 0),
            frontend_handled=frontend_status.get("handled", 0),
            frontend_requests_total=frontend_status.get("requests_total", 0),
            # DB metrics
            db_connections=mysql_stats.get("Threads_connected", 0),
            db_threads_running=mysql_stats.get("Threads_running", 0),
            db_errors=db_errors,
            db_total_queries=mysql_stats.get("Questions", 0),
        )
        
        # Calculate backend memory
        usage = resource.getrusage(resource.RUSAGE_SELF)
        backend_memory_mb = float(usage.ru_maxrss) / 1024.0
        if backend_memory_mb > 1024 * 1024:
            backend_memory_mb = float(usage.ru_maxrss) / (1024.0 * 1024.0)
        metric.backend_memory_mb = round(backend_memory_mb, 2)
        
        db.add(metric)
        db.commit()
        db.close()
    except Exception as e:
        print(f"Error recording insights snapshot: {e}")


def background_insights_recorder():
    """Background task that records insights metrics every 30 seconds."""
    global insights_recording_active
    while insights_recording_active:
        try:
            record_insights_snapshot()
        except Exception as e:
            print(f"Background insights recording error: {e}")
        time.sleep(30)  # Record every 30 seconds


@app.on_event("startup")
def startup_insights_recording():
    """Start the background insights recording task on app startup."""
    ensure_superadmin_account()
    global insights_recording_thread, insights_recording_active
    if not insights_recording_active:
        insights_recording_active = True
        insights_recording_thread = threading.Thread(target=background_insights_recorder, daemon=True)
        insights_recording_thread.start()
        print("App Insights historical recording started (30s interval)")


@app.on_event("shutdown")
def shutdown_insights_recording():
    """Stop the background insights recording task on app shutdown."""
    global insights_recording_active
    insights_recording_active = False
    print("App Insights historical recording stopped")


# ==================== Health Check ====================

@app.get("/health")
def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "timestamp": datetime.utcnow()}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
